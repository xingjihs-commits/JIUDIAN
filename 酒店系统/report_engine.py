"""
======================================================
ShadowGuard — 报表与数据分析引擎 (v1.0)

功能：
  - 月度/年度营收报表
  - 出租率趋势图（纯 QPainter，无需 matplotlib）
  - 导出 Excel（openpyxl）/ CSV
  - 房型收益对比分析
  - 报表标签页（嵌入 WorkspaceDock）

使用入口：WorkspaceDock → 📊 报表标签页
======================================================
"""
from __future__ import annotations

import csv
import io
import os
from datetime import datetime, timedelta, date
from typing import Any, Optional

from PySide6.QtCore import Qt, QRect, QTimer, QThread, Signal as QtSignal
from PySide6.QtGui import (
    QColor, QBrush, QPainter, QPen, QFont, QFontMetrics,
    QPaintEvent, QLinearGradient
)
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QDialog, QFileDialog, QTabWidget, QFrame,
    QScrollArea, QSizePolicy, QProgressBar, QSplitter,
    QGridLayout, QGroupBox
)

from database import db, LEDGER_REVENUE_TX_TYPES
from ui_helpers import show_info, show_warning, show_error, ask_confirm
from design_tokens import _p
from theme_palette import _lighten_hex, _darken_hex

# ─────────────────────────────────────────────
#  颜色常量（跟随当前主题）
# ─────────────────────────────────────────────
def _build_chart_colors() -> tuple[list[QColor], QColor, QColor, QColor, QColor, QColor]:
    from design_tokens import _p
    bg = _p("bg")
    grid = _p("border")
    axis = _p("text_muted")
    text = _p("text")
    accent = _p("primary")
    primary = _p("primary")
    amount_pos = _p("amount_positive")
    danger_c = _p("danger")
    gold = _p("accent")
    return [
        QColor(primary), QColor(amount_pos), QColor(gold),
        QColor(danger_c), QColor(_lighten_hex(primary, 0.3)), QColor(_darken_hex(amount_pos, 0.2)),
        QColor(_lighten_hex(gold, 0.2)), QColor(_darken_hex(danger_c, 0.2)),
    ], QColor(bg), QColor(grid), QColor(axis), QColor(text), QColor(accent)

# 图表主题色在 paintEvent 内通过 _build_chart_colors() 实时读取，换主题后 update() 即可生效。


# ─────────────────────────────────────────────
#  数据查询层
# ─────────────────────────────────────────────

class ReportData:
    """报表数据查询（所有查询集中在此类）"""

    @staticmethod
    def monthly_revenue(year: int, month: int) -> dict[str, Any]:
        """月度营收汇总"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"

        # 收入口径 = 营业额（LEDGER_REVENUE_TX_TYPES），不含押金；与 database / 财务标签页一致。
        # 支出只统计 PAYOUT，排除 SHIFT_DIFF/NIGHT_AUDIT 等非现金流类型。
        inc_in = ",".join(f"'{t}'" for t in LEDGER_REVENUE_TX_TYPES)
        rows = db.execute(
            f"""SELECT DATE(created_at) as day,
                      SUM(CASE WHEN tx_type IN ({inc_in}) THEN amount ELSE 0 END) as income,
                      SUM(CASE WHEN tx_type='PAYOUT' THEN ABS(amount) ELSE 0 END) as expense
               FROM ledger
               WHERE created_at >= ? AND created_at < ?
               GROUP BY day ORDER BY day""",
            (start, end)
        ).fetchall()

        days, incomes, expenses = [], [], []
        for r in rows:
            days.append(r[0])
            incomes.append(float(r[1] or 0))
            expenses.append(float(r[2] or 0))

        total_income  = sum(incomes)
        total_expense = sum(expenses)
        return {
            "days": days,
            "incomes": incomes,
            "expenses": expenses,
            "total_income": total_income,
            "total_expense": total_expense,
            "net_profit": total_income - total_expense,
        }

    @staticmethod
    def yearly_revenue(year: int) -> dict[str, Any]:
        """年度月度营收汇总"""
        months = list(range(1, 13))
        incomes, expenses = [], []
        for m in months:
            data = ReportData.monthly_revenue(year, m)
            incomes.append(data["total_income"])
            expenses.append(data["total_expense"])
        return {
            "months": [f"{m}月" for m in months],
            "incomes": incomes,
            "expenses": expenses,
            "net_profits": [i - e for i, e in zip(incomes, expenses)],
            "total_income": sum(incomes),
            "total_expense": sum(expenses),
            "net_profit": sum(incomes) - sum(expenses),
        }

    @staticmethod
    def occupancy_trend(days: int = 30) -> dict[str, Any]:
        """出租率趋势（最近N天）
        
        修复：历史出租率不能用 status='INHOUSE' 条件（已退房客人状态已变为 CHECKOUT）。
        改为查询 guests 表的入住/退房时间范围，判断某天是否有客人在住。
        """
        total_rooms_row = db.execute("SELECT COUNT(*) FROM rooms").fetchone()
        total_rooms = total_rooms_row[0] if total_rooms_row else 1
        if total_rooms == 0:
            total_rooms = 1

        result_days, rates = [], []
        today = date.today()
        for i in range(days - 1, -1, -1):
            d = today - timedelta(days=i)
            d_str = d.strftime("%Y-%m-%d")
            # 修复：不再过滤 status='INHOUSE'，而是通过时间范围判断当天在住情况
            # 条件：入住时间 <= 当天23:59:59 AND (退房时间 IS NULL OR 退房时间 > 当天00:00:00)
            row = db.execute(
                """SELECT COUNT(DISTINCT room_id) FROM guests
                   WHERE checkin_time <= ?
                   AND (checkout_time IS NULL OR checkout_time > ?)""",
                (d_str + " 23:59:59", d_str + " 00:00:00")
            ).fetchone()
            occupied = row[0] if row else 0
            rate = round(occupied / total_rooms * 100, 1)
            result_days.append(d.strftime("%m/%d"))
            rates.append(rate)

        avg_rate = round(sum(rates) / len(rates), 1) if rates else 0
        return {
            "days": result_days,
            "rates": rates,
            "avg_rate": avg_rate,
            "total_rooms": total_rooms,
        }

    @staticmethod
    def room_type_revenue(year: int, month: int) -> dict[str, Any]:
        """房型收益对比"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"

        rev_in = ",".join(f"'{t}'" for t in LEDGER_REVENUE_TX_TYPES)
        rows = db.execute(
            f"""SELECT r.room_type,
                      COUNT(DISTINCT g.id) as checkins,
                      SUM(l.amount) as revenue
               FROM rooms r
               LEFT JOIN guests g ON g.room_id = r.room_id
                   AND g.checkin_time >= ? AND g.checkin_time < ?
               LEFT JOIN ledger l ON l.room_id = r.room_id
                   AND l.created_at >= ? AND l.created_at < ?
                   AND l.amount > 0
                   AND l.tx_type IN ({rev_in})
               GROUP BY r.room_type""",
            (start, end, start, end)
        ).fetchall()

        types, checkins, revenues = [], [], []
        for r in rows:
            types.append(r[0] or "未知")
            checkins.append(int(r[1] or 0))
            revenues.append(float(r[2] or 0))

        return {
            "types": types,
            "checkins": checkins,
            "revenues": revenues,
        }

    @staticmethod
    def payment_method_breakdown(year: int, month: int) -> dict[str, Any]:
        """支付方式分布"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"

        rows = db.execute(
            """SELECT pay_method, SUM(amount) as total
               FROM ledger
               WHERE amount > 0 AND created_at >= ? AND created_at < ?
               GROUP BY pay_method ORDER BY total DESC""",
            (start, end)
        ).fetchall()

        methods = [r[0] or "未知" for r in rows]
        amounts = [float(r[1] or 0) for r in rows]
        total = sum(amounts)
        percentages = [round(a / total * 100, 1) if total > 0 else 0 for a in amounts]

        return {
            "methods": methods,
            "amounts": amounts,
            "percentages": percentages,
            "total": total,
        }

    @staticmethod
    def top_guests(limit: int = 10) -> list[dict]:
        """消费最多的住客
        
        修复：原查询通过 room_id 关联账单，导致同一房间所有历史住客共享全部账单金额。
        改为通过 guest 的入住/退房时间范围关联账单，确保账单归属正确。
        """
        rev_in = ",".join(f"'{t}'" for t in LEDGER_REVENUE_TX_TYPES)
        rows = db.execute(
            f"""SELECT g.name, g.phone, COUNT(DISTINCT g.id) as stays,
                      COALESCE(SUM(l.amount), 0) as total_spend
               FROM guests g
               LEFT JOIN ledger l
                   ON l.room_id = g.room_id
                   AND l.amount > 0
                   AND l.tx_type IN ({rev_in})
                   AND l.created_at >= g.checkin_time
                   AND (g.checkout_time IS NULL OR l.created_at <= g.checkout_time)
               WHERE g.name IS NOT NULL AND g.name != ''
               GROUP BY g.name, g.phone
               ORDER BY total_spend DESC LIMIT ?""",
            (limit,)
        ).fetchall()
        return [
            {"name": r[0], "phone": r[1] or "", "stays": r[2], "total": float(r[3] or 0)}
            for r in rows
        ]

    # —— Round 3.2 新增报表 ——

    @staticmethod
    def channel_analysis(year: int, month: int) -> dict[str, Any]:
        """客源渠道分析（OTA/散客/协议/会员占比）。"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"
        rows = db.execute(
            "SELECT COALESCE(channel, 'WALK_IN'), COUNT(*), COALESCE(SUM(amount),0) "
            "FROM ledger WHERE created_at>=? AND created_at<? AND tx_type IN ('ROOM_IN','SHOP') "
            "GROUP BY channel",
            (start, end),
        ).fetchall()
        return {
            "period": f"{year}-{month:02d}",
            "channels": [{"channel": r[0], "count": r[1], "revenue": float(r[2] or 0)} for r in rows],
        }

    @staticmethod
    def product_sales_ranking(year: int, month: int, limit: int = 10) -> list[dict]:
        """商品销量排行榜 & 毛利率。"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"
        rows = db.execute(
            "SELECT l.note, COUNT(*), COALESCE(SUM(l.amount),0), "
            "COALESCE(s.cost_price,0), COALESCE(s.price,0) "
            "FROM ledger l LEFT JOIN shop_items s ON s.name=l.note "
            "WHERE l.created_at>=? AND l.created_at<? AND l.tx_type='SHOP' AND l.note IS NOT NULL "
            "GROUP BY l.note ORDER BY COUNT(*) DESC LIMIT ?",
            (start, end, limit),
        ).fetchall()
        return [
            {
                "item": r[0],
                "sales_count": r[1],
                "revenue": float(r[2] or 0),
                "cost_per_unit": float(r[3] or 0),
                "price": float(r[4] or 0),
                "margin_pct": round((float(r[4] or 0) - float(r[3] or 0)) / max(float(r[4] or 1), 0.01) * 100, 1),
            }
            for r in rows
        ]

    @staticmethod
    def daily_energy_report(date_str: str) -> dict[str, Any]:
        """能耗日报（水电气对比）。"""
        rows = db.execute(
            "SELECT meter_id, reading_type, COALESCE(SUM(reading_value),0) "
            "FROM energy_meter_readings WHERE date(created_at)=? "
            "GROUP BY meter_id, reading_type",
            (date_str,),
        ).fetchall()
        return {
            "date": date_str,
            "readings": [{"meter": r[0], "type": r[1], "total": float(r[2] or 0)} for r in rows],
        }

    @staticmethod
    def energy_anomaly_top10(limit: int = 10) -> list[dict]:
        """能耗异常房间 TOP10。"""
        rows = db.execute(
            "SELECT room_id, reading_type, reading_value, created_at, is_anomaly "
            "FROM energy_audit WHERE is_anomaly=1 "
            "ORDER BY created_at DESC LIMIT ?",
            (limit,),
        ).fetchall()
        return [{"room": r[0], "type": r[1], "value": float(r[2] or 0), "time": str(r[3])} for r in rows]

    @staticmethod
    def staff_performance(start: str, end: str) -> list[dict]:
        """员工绩效（入住数/退房数/收款额）。"""
        rows = db.execute(
            "SELECT operator_id, tx_type, COUNT(*) as cnt, COALESCE(SUM(amount),0) as total "
            "FROM ledger WHERE created_at BETWEEN ? AND ? "
            "AND tx_type IN ('ROOM_IN','CHECKOUT','SHOP','NIGHT_AUDIT') "
            "GROUP BY operator_id, tx_type ORDER BY operator_id",
            (start, end),
        ).fetchall()
        perf = {}
        for r in rows:
            if r[0] not in perf:
                perf[r[0]] = {"staff": r[0], "checkins": 0, "checkouts": 0, "revenue": 0.0, "night_audits": 0}
            if r[1] == "ROOM_IN":
                perf[r[0]]["checkins"] += r[2]
            elif r[1] == "CHECKOUT":
                perf[r[0]]["checkouts"] += r[2]
            elif r[1] == "SHOP":
                perf[r[0]]["revenue"] += float(r[3] or 0)
            elif r[1] == "NIGHT_AUDIT":
                perf[r[0]]["night_audits"] += r[2]
        return sorted(perf.values(), key=lambda x: x["revenue"], reverse=True)

    @staticmethod
    def member_spending_analysis(year: int, month: int) -> list[dict]:
        """会员消费分析。"""
        start = f"{year:04d}-{month:02d}-01"
        if month == 12:
            end = f"{year+1:04d}-01-01"
        else:
            end = f"{year:04d}-{month+1:02d}-01"
        rows = db.execute(
            "SELECT m.name, m.phone, m.grade, COUNT(*), COALESCE(SUM(l.amount),0) "
            "FROM ledger l JOIN members m ON l.guest_phone=m.phone "
            "WHERE l.created_at>=? AND l.created_at<? "
            "GROUP BY m.id ORDER BY SUM(l.amount) DESC LIMIT 20",
            (start, end),
        ).fetchall()
        return [
            {"name": r[0], "phone": r[1], "grade": r[2], "visits": r[3], "total_spent": float(r[4] or 0)}
            for r in rows
        ]

    @staticmethod
    def deposit_flow_report(start: str, end: str) -> dict[str, Any]:
        """押金流水报告。"""
        rows = db.execute(
            "SELECT DATE(created_at), tx_type, COALESCE(SUM(amount),0), COUNT(*) "
            "FROM ledger WHERE created_at BETWEEN ? AND ? "
            "AND tx_type in ('DEPOSIT','DEPOSIT_REFUND','DEPOSIT_DEDUCT') "
            "GROUP BY DATE(created_at), tx_type ORDER BY DATE(created_at)",
            (start, end),
        ).fetchall()
        return {
            "period": f"{start} ~ {end}",
            "entries": [
                {"date": r[0], "type": r[1], "total": float(r[2] or 0), "count": r[3]} for r in rows
            ],
        }

    @staticmethod
    def multi_currency_summary(year: int, month: int) -> dict[str, Any]:
        """[sub-a] 多币种财务汇总：按 ledger.currency 分组 + 本位币合计列。

        业务背景：原 monthly_revenue 直接 SUM(amount) 跨币种相加，外币收款被
        当本位币计算，月报与银行对账单差异无法解释。

        本方法：
          1. 按 ledger.currency 分组 SUM(amount) 得到各币种原币小计
          2. 按 AVG(exchange_rate) 折算到本位币（money_utils.Decimal）
          3. 输出按币种明细 + 本位币合计列
          4. 与单币种 monthly_revenue 互不依赖，可并存

        所有金额运算走 Decimal，避免 float 累计误差。
        """
        from money_utils import base_currency, quantize_money, to_base, to_money
        from decimal import Decimal as _D

        start = f"{year:04d}-{month:02d}-01"
        end = f"{year+1:04d}-01-01" if month == 12 else f"{year:04d}-{month+1:02d}-01"
        base = base_currency()

        # 收入流水按币种分组（amount > 0 排除退款冲抵行）
        rows = db.execute(
            """
            SELECT COALESCE(currency, ?) AS cur,
                   COALESCE(SUM(amount), 0) AS orig_total,
                   AVG(COALESCE(exchange_rate, 1.0)) AS avg_rate
            FROM ledger
            WHERE created_at >= ? AND created_at < ?
              AND amount > 0
            GROUP BY cur
            ORDER BY cur
            """,
            (base, start, end),
        ).fetchall()

        by_currency: list[dict[str, Any]] = []
        grand_total_base = _D("0")
        for cur, orig_total, avg_rate in rows:
            currency = (cur or base).upper()
            orig_dec = to_money(orig_total)
            rate = to_money(avg_rate) or _D("1")
            base_amt = to_base(orig_dec, currency, rate)
            by_currency.append({
                "currency": currency,
                "orig_total": float(quantize_money(orig_dec)),
                "avg_rate": float(quantize_money(rate)),
                "base_total": float(quantize_money(base_amt)),
            })
            grand_total_base += base_amt

        return {
            "period": f"{year}-{month:02d}",
            "base_currency": base,
            "by_currency": by_currency,
            "grand_total_base": float(quantize_money(grand_total_base)),
        }


# ─────────────────────────────────────────────
#  折线图组件
# ─────────────────────────────────────────────

class LineChartWidget(QWidget):
    """纯 QPainter 折线图（支持多条线）"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._title = ""
        self._labels: list[str] = []
        self._series: list[dict] = []  # [{"name": str, "data": list[float], "color": QColor}]
        self._y_unit = ""
        self.setMinimumHeight(220)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, title: str, labels: list[str],
                 series: list[dict], y_unit: str = "") -> None:
        self._title = title
        self._labels = labels
        self._series = series
        self._y_unit = y_unit
        self.update()

    def paintEvent(self, event: QPaintEvent):
        if not self._labels or not self._series:
            return
        chart_colors, color_bg, color_grid, color_axis, color_text, _color_accent = (
            _build_chart_colors()
        )
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        W, H = self.width(), self.height()
        PAD_L, PAD_R, PAD_T, PAD_B = 60, 20, 30, 40

        chart_w = W - PAD_L - PAD_R
        chart_h = H - PAD_T - PAD_B

        # 背景
        p.fillRect(0, 0, W, H, color_bg)

        # 标题
        font_title = QFont()
        font_title.setPointSize(11)
        font_title.setBold(True)
        p.setFont(font_title)
        p.setPen(color_text)
        p.drawText(QRect(PAD_L, 4, chart_w, PAD_T - 4), Qt.AlignCenter, self._title)

        # 计算数据范围
        all_vals = [v for s in self._series for v in s["data"]]
        if not all_vals:
            return
        y_min = 0
        y_max = max(all_vals) * 1.1 if max(all_vals) > 0 else 100

        # 网格线
        p.setPen(QPen(color_grid, 1, Qt.DashLine))
        font_small = QFont()
        font_small.setPointSize(8)
        p.setFont(font_small)
        p.setPen(color_axis)

        grid_count = 5
        for i in range(grid_count + 1):
            y_val = y_min + (y_max - y_min) * i / grid_count
            y_px = PAD_T + chart_h - int(chart_h * i / grid_count)
            p.setPen(QPen(color_grid, 1, Qt.DashLine))
            p.drawLine(PAD_L, y_px, PAD_L + chart_w, y_px)
            p.setPen(color_axis)
            label = f"{y_val:.0f}{self._y_unit}"
            p.drawText(QRect(0, y_px - 8, PAD_L - 4, 16), Qt.AlignRight | Qt.AlignVCenter, label)

        # X 轴标签（每隔几个显示一个）
        n = len(self._labels)
        step = max(1, n // 10)
        for i, lbl in enumerate(self._labels):
            if i % step != 0 and i != n - 1:
                continue
            x_px = PAD_L + int(chart_w * i / max(n - 1, 1))
            p.setPen(color_axis)
            p.drawText(QRect(x_px - 20, PAD_T + chart_h + 4, 40, 16),
                       Qt.AlignCenter, lbl)

        # 绘制折线
        for si, series in enumerate(self._series):
            color = series.get("color", chart_colors[si % len(chart_colors)])
            data = series["data"]
            if not data:
                continue

            points = []
            for i, val in enumerate(data):
                x_px = PAD_L + int(chart_w * i / max(len(data) - 1, 1))
                y_px = PAD_T + chart_h - int(chart_h * (val - y_min) / max(y_max - y_min, 1))
                points.append((x_px, y_px))

            # 填充区域（半透明）
            fill_color = QColor(color)
            fill_color.setAlpha(30)
            p.setBrush(QBrush(fill_color))
            p.setPen(Qt.NoPen)
            poly_points = [(PAD_L, PAD_T + chart_h)] + points + [(PAD_L + chart_w, PAD_T + chart_h)]
            from PySide6.QtGui import QPolygon
            from PySide6.QtCore import QPoint
            poly = QPolygon([QPoint(x, y) for x, y in poly_points])
            p.drawPolygon(poly)

            # 折线
            p.setBrush(Qt.NoBrush)
            p.setPen(QPen(color, 2))
            for i in range(len(points) - 1):
                p.drawLine(points[i][0], points[i][1], points[i+1][0], points[i+1][1])

            # 数据点
            p.setBrush(QBrush(color))
            for x_px, y_px in points:
                p.drawEllipse(x_px - 3, y_px - 3, 6, 6)

        # 图例
        legend_x = PAD_L
        for si, series in enumerate(self._series):
            color = series.get("color", chart_colors[si % len(chart_colors)])
            p.setBrush(QBrush(color))
            p.setPen(Qt.NoPen)
            p.drawRect(legend_x, 8, 12, 12)
            p.setPen(color_text)
            p.setFont(font_small)
            p.drawText(legend_x + 16, 8, 80, 12, Qt.AlignVCenter, series["name"])
            legend_x += 100

        p.end()


# ─────────────────────────────────────────────
#  柱状图组件
# ─────────────────────────────────────────────

class BarChartWidget(QWidget):
    """纯 QPainter 柱状图"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._title = ""
        self._labels: list[str] = []
        self._series: list[dict] = []
        self._y_unit = ""
        self.setMinimumHeight(220)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, title: str, labels: list[str],
                 series: list[dict], y_unit: str = "") -> None:
        self._title = title
        self._labels = labels
        self._series = series
        self._y_unit = y_unit
        self.update()

    def paintEvent(self, event: QPaintEvent):
        if not self._labels or not self._series:
            return
        chart_colors, color_bg, color_grid, color_axis, color_text, _color_accent = (
            _build_chart_colors()
        )
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        W, H = self.width(), self.height()
        PAD_L, PAD_R, PAD_T, PAD_B = 70, 20, 30, 50

        chart_w = W - PAD_L - PAD_R
        chart_h = H - PAD_T - PAD_B

        p.fillRect(0, 0, W, H, color_bg)

        # 标题
        font_title = QFont()
        font_title.setPointSize(11)
        font_title.setBold(True)
        p.setFont(font_title)
        p.setPen(color_text)
        p.drawText(QRect(PAD_L, 4, chart_w, PAD_T - 4), Qt.AlignCenter, self._title)

        all_vals = [v for s in self._series for v in s["data"]]
        if not all_vals:
            return
        y_max = max(all_vals) * 1.1 if max(all_vals) > 0 else 100

        # 网格
        font_small = QFont()
        font_small.setPointSize(8)
        p.setFont(font_small)
        grid_count = 5
        for i in range(grid_count + 1):
            y_val = y_max * i / grid_count
            y_px = PAD_T + chart_h - int(chart_h * i / grid_count)
            p.setPen(QPen(color_grid, 1, Qt.DashLine))
            p.drawLine(PAD_L, y_px, PAD_L + chart_w, y_px)
            p.setPen(color_axis)
            label = f"{y_val:.0f}{self._y_unit}"
            p.drawText(QRect(0, y_px - 8, PAD_L - 4, 16), Qt.AlignRight | Qt.AlignVCenter, label)

        # 柱子
        n = len(self._labels)
        ns = len(self._series)
        group_w = chart_w / max(n, 1)
        bar_w = group_w * 0.7 / max(ns, 1)

        for si, series in enumerate(self._series):
            color = series.get("color", chart_colors[si % len(chart_colors)])
            for i, val in enumerate(series["data"]):
                bar_h = int(chart_h * val / max(y_max, 1))
                x = PAD_L + int(group_w * i + group_w * 0.15 + bar_w * si)
                y = PAD_T + chart_h - bar_h

                # 渐变
                grad = QLinearGradient(x, y, x, y + bar_h)
                grad.setColorAt(0, color.lighter(120))
                grad.setColorAt(1, color)
                p.setBrush(QBrush(grad))
                p.setPen(Qt.NoPen)
                p.drawRoundedRect(int(x), y, int(bar_w), bar_h, 3, 3)

                # 数值标签
                if bar_h > 20:
                    p.setPen(QColor(_p("surface")))
                    p.setFont(font_small)
                    p.drawText(QRect(int(x), y + 2, int(bar_w), 14),
                               Qt.AlignCenter, f"{val:.0f}")

        # X 轴标签
        p.setPen(color_axis)
        p.setFont(font_small)
        for i, lbl in enumerate(self._labels):
            x = PAD_L + int(group_w * i + group_w * 0.5)
            p.drawText(QRect(x - 25, PAD_T + chart_h + 4, 50, 20),
                       Qt.AlignCenter, lbl)

        # 图例
        legend_x = PAD_L
        for si, series in enumerate(self._series):
            color = series.get("color", chart_colors[si % len(chart_colors)])
            p.setBrush(QBrush(color))
            p.setPen(Qt.NoPen)
            p.drawRect(legend_x, 8, 12, 12)
            p.setPen(color_text)
            p.setFont(font_small)
            p.drawText(legend_x + 16, 8, 80, 12, Qt.AlignVCenter, series["name"])
            legend_x += 100

        p.end()


# ─────────────────────────────────────────────
#  饼图组件
# ─────────────────────────────────────────────

class PieChartWidget(QWidget):
    """纯 QPainter 饼图"""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self._title = ""
        self._labels: list[str] = []
        self._values: list[float] = []
        self.setMinimumHeight(200)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, title: str, labels: list[str], values: list[float]) -> None:
        self._title = title
        self._labels = labels
        self._values = values
        self.update()

    def paintEvent(self, event: QPaintEvent):
        if not self._values:
            return
        chart_colors, color_bg, _color_grid, _color_axis, color_text, _color_accent = (
            _build_chart_colors()
        )
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        W, H = self.width(), self.height()
        p.fillRect(0, 0, W, H, color_bg)

        # 标题
        font_title = QFont()
        font_title.setPointSize(11)
        font_title.setBold(True)
        p.setFont(font_title)
        p.setPen(color_text)
        p.drawText(QRect(0, 4, W, 24), Qt.AlignCenter, self._title)

        total = sum(self._values)
        if total == 0:
            return

        # 饼图区域
        pie_size = min(W - 160, H - 60)
        pie_x = 20
        pie_y = 30 + (H - 60 - pie_size) // 2

        start_angle = 0
        font_small = QFont()
        font_small.setPointSize(8)

        for i, (lbl, val) in enumerate(zip(self._labels, self._values)):
            color = chart_colors[i % len(chart_colors)]
            span = int(val / total * 360 * 16)

            p.setBrush(QBrush(color))
            p.setPen(QPen(QColor(_p("surface")), 2))
            p.drawPie(pie_x, pie_y, pie_size, pie_size, start_angle, span)

            start_angle += span

        # 图例
        legend_x = pie_x + pie_size + 20
        legend_y = 40
        p.setFont(font_small)
        for i, (lbl, val) in enumerate(zip(self._labels, self._values)):
            color = chart_colors[i % len(chart_colors)]
            pct = val / total * 100
            p.setBrush(QBrush(color))
            p.setPen(Qt.NoPen)
            p.drawRect(legend_x, legend_y + i * 22, 12, 12)
            p.setPen(color_text)
            p.drawText(legend_x + 16, legend_y + i * 22, 120, 14,
                       Qt.AlignVCenter, f"{lbl} {pct:.1f}%")

        p.end()


# ─────────────────────────────────────────────
#  KPI 卡片组件
# ─────────────────────────────────────────────

class KpiCard(QFrame):
    """KPI 数字卡片"""

    def __init__(self, title: str, value: str, subtitle: str = "",
                 tone: str = "primary", parent=None):
        super().__init__(parent)
        self._tone = tone
        self.setObjectName("ReportKpiCard")
        self.setFrameShape(QFrame.StyledPanel)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(4)

        self._title_lbl = QLabel(title)
        self._title_lbl.setObjectName("ReportKpiTitle")
        layout.addWidget(self._title_lbl)

        self.value_lbl = QLabel(value)
        self.value_lbl.setObjectName("ReportKpiValue")
        self.value_lbl.setProperty("tone", tone)
        layout.addWidget(self.value_lbl)

        self._sub_lbl = None
        if subtitle:
            self._sub_lbl = QLabel(subtitle)
            self._sub_lbl.setObjectName("ReportKpiSub")
            layout.addWidget(self._sub_lbl)

    def refresh_theme(self) -> None:
        self.value_lbl.setProperty("tone", self._tone)
        self.value_lbl.style().unpolish(self.value_lbl)
        self.value_lbl.style().polish(self.value_lbl)

    def update_value(self, value: str, subtitle: str = "") -> None:
        self.value_lbl.setText(value)


# ─────────────────────────────────────────────
#  导出功能
# ─────────────────────────────────────────────

class ReportExporter:
    """报表导出（Excel / CSV）"""

    @staticmethod
    def export_monthly_csv(year: int, month: int, filepath: str) -> tuple[bool, str]:
        """导出月度报表为 CSV"""
        try:
            data = ReportData.monthly_revenue(year, month)
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([f"{year}年{month}月 营收报表"])
                writer.writerow([])
                writer.writerow(["日期", "收入", "支出", "净利润"])
                for i, day in enumerate(data["days"]):
                    income = data["incomes"][i]
                    expense = data["expenses"][i]
                    writer.writerow([day, f"{income:.2f}", f"{expense:.2f}",
                                     f"{income - expense:.2f}"])
                writer.writerow([])
                writer.writerow(["合计",
                                  f"{data['total_income']:.2f}",
                                  f"{data['total_expense']:.2f}",
                                  f"{data['net_profit']:.2f}"])
            return True, filepath
        except Exception as e:
            return False, str(e)

    @staticmethod
    def export_monthly_excel(year: int, month: int, filepath: str) -> tuple[bool, str]:
        """导出月度报表为 Excel（需要 openpyxl）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{year}年{month}月报表"

            # 标题样式
            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center = Alignment(horizontal="center", vertical="center")

            # 标题行
            ws.merge_cells("A1:D1")
            ws["A1"] = f"{year}年{month}月 营收报表"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = center

            # 表头
            headers = ["日期", "收入", "支出", "净利润"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            # 数据
            data = ReportData.monthly_revenue(year, month)
            for i, day in enumerate(data["days"]):
                row = i + 3
                income = data["incomes"][i]
                expense = data["expenses"][i]
                net = income - expense
                ws.cell(row=row, column=1, value=day)
                ws.cell(row=row, column=2, value=round(income, 2))
                ws.cell(row=row, column=3, value=round(expense, 2))
                cell_net = ws.cell(row=row, column=4, value=round(net, 2))
                if net < 0:
                    cell_net.font = Font(color="DC2626")

            # 合计行
            last_row = len(data["days"]) + 3
            ws.cell(row=last_row, column=1, value="合计").font = Font(bold=True)
            ws.cell(row=last_row, column=2, value=round(data["total_income"], 2)).font = Font(bold=True)
            ws.cell(row=last_row, column=3, value=round(data["total_expense"], 2)).font = Font(bold=True)
            ws.cell(row=last_row, column=4, value=round(data["net_profit"], 2)).font = Font(bold=True)

            # 出租率工作表
            ws2 = wb.create_sheet("出租率趋势")
            occ = ReportData.occupancy_trend(30)
            ws2["A1"] = "日期"
            ws2["B1"] = "出租率(%)"
            for i, (d, r) in enumerate(zip(occ["days"], occ["rates"])):
                ws2.cell(row=i+2, column=1, value=d)
                ws2.cell(row=i+2, column=2, value=r)

            # 房型收益工作表
            ws3 = wb.create_sheet("房型收益")
            rt = ReportData.room_type_revenue(year, month)
            ws3["A1"] = "房型"
            ws3["B1"] = "入住次数"
            ws3["C1"] = "营收"
            for i, (t, c, r) in enumerate(zip(rt["types"], rt["checkins"], rt["revenues"])):
                ws3.cell(row=i+2, column=1, value=t)
                ws3.cell(row=i+2, column=2, value=c)
                ws3.cell(row=i+2, column=3, value=round(r, 2))

            # 调整列宽
            for sheet in [ws, ws2, ws3]:
                for col in sheet.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    sheet.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

            wb.save(filepath)
            return True, filepath
        except ImportError:
            # openpyxl 未安装，降级为 CSV
            csv_path = filepath.replace(".xlsx", ".csv")
            return ReportExporter.export_monthly_csv(year, month, csv_path)
        except Exception as e:
            return False, str(e)

    @staticmethod
    def export_pdf(year: int, month: int, filepath: str) -> tuple[bool, str]:
        """
        导出月度报表为 PDF。
        优先使用 reportlab；若未安装则降级为纯文本 .txt 文件。
        返回 (success: bool, filepath_or_error: str)
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph,
                Spacer, HRFlowable
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os as _os

            # ── 尝试注册中文字体（优先系统字体）──
            font_name = "Helvetica"  # 默认回退
            import deploy_paths
            _fonts = deploy_paths.fonts_dir()
            font_candidates = [
                # Windows
                _os.path.join(_fonts, "msyh.ttc"),
                _os.path.join(_fonts, "simhei.ttf"),
                _os.path.join(_fonts, "simsun.ttc"),
                # macOS
                "/System/Library/Fonts/PingFang.ttc",
                "/Library/Fonts/Arial Unicode MS.ttf",
                # Linux
                "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            ]
            for fp in font_candidates:
                if _os.path.exists(fp):
                    try:
                        pdfmetrics.registerFont(TTFont("CJK", fp))
                        font_name = "CJK"
                        break
                    except Exception:
                        continue

            # ── 数据采集 ──
            cur  = ReportData.monthly_revenue(year, month)
            # 上月
            prev_month = month - 1 if month > 1 else 12
            prev_year  = year if month > 1 else year - 1
            prev = ReportData.monthly_revenue(prev_year, prev_month)

            def _pct_change(a, b):
                return ((a - b) / b * 100) if b else 0.0

            # 入住次数
            start = f"{year:04d}-{month:02d}-01"
            end   = f"{year:04d}-{month+1:02d}-01" if month < 12 else f"{year+1:04d}-01-01"
            checkins_row = db.execute(
                "SELECT COUNT(*) FROM guests WHERE checkin_time >= ? AND checkin_time < ?",
                (start, end)
            ).fetchone()
            checkins = checkins_row[0] if checkins_row else 0

            prev_start = f"{prev_year:04d}-{prev_month:02d}-01"
            prev_end   = f"{prev_year:04d}-{prev_month+1:02d}-01" if prev_month < 12 else f"{prev_year+1:04d}-01-01"
            prev_checkins_row = db.execute(
                "SELECT COUNT(*) FROM guests WHERE checkin_time >= ? AND checkin_time < ?",
                (prev_start, prev_end)
            ).fetchone()
            prev_checkins = prev_checkins_row[0] if prev_checkins_row else 0

            occ  = ReportData.occupancy_trend(30)
            rt   = ReportData.room_type_revenue(year, month)
            top  = ReportData.top_guests(10)
            hotel    = db.get_config("hotel_name") or "酒店"
            currency = db.get_config("currency_symbol") or "¥"

            # 构建 summary 兼容字典
            summary = {
                "income":         cur["total_income"],
                "expense":        cur["total_expense"],
                "profit":         cur["net_profit"],
                "prev_income":    prev["total_income"],
                "prev_expense":   prev["total_expense"],
                "prev_profit":    prev["net_profit"],
                "income_change":  _pct_change(cur["total_income"],  prev["total_income"]),
                "expense_change": _pct_change(cur["total_expense"], prev["total_expense"]),
                "profit_change":  _pct_change(cur["net_profit"],    prev["net_profit"]),
                "checkins":       checkins,
                "prev_checkins":  prev_checkins,
                "checkin_change": _pct_change(checkins, prev_checkins),
                "avg_occ":        occ["avg_rate"],
                "prev_avg_occ":   0.0,  # 历史出租率暂不跨月计算
            }

            # ── 样式 ──
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "Title", fontName=font_name, fontSize=18,
                textColor=colors.HexColor(_p("primary")),
                spaceAfter=6, alignment=1  # 居中
            )
            h2_style = ParagraphStyle(
                "H2", fontName=font_name, fontSize=13,
                textColor=colors.HexColor(_p("primary")),
                spaceBefore=12, spaceAfter=4
            )
            body_style = ParagraphStyle(
                "Body", fontName=font_name, fontSize=10,
                textColor=colors.HexColor(_p("text")), spaceAfter=2
            )

            # ── 构建内容 ──
            story = []
            page_w, page_h = A4

            # 标题
            story.append(Paragraph(
                f"{hotel} · {year}年{month}月 运营报表",
                title_style
            ))
            story.append(Paragraph(
                f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
                body_style
            ))
            story.append(HRFlowable(width="100%", thickness=1,
                                     color=colors.HexColor(_p("border"))))
            story.append(Spacer(1, 6*mm))

            # KPI 汇总表
            story.append(Paragraph("📊 月度核心指标", h2_style))
            kpi_data = [
                ["指标", "本月", "上月", "环比"],
                ["总收入",
                 f"{currency}{summary['income']:,.2f}",
                 f"{currency}{summary['prev_income']:,.2f}",
                 f"{summary['income_change']:+.1f}%"],
                ["总支出",
                 f"{currency}{summary['expense']:,.2f}",
                 f"{currency}{summary['prev_expense']:,.2f}",
                 f"{summary['expense_change']:+.1f}%"],
                ["净利润",
                 f"{currency}{summary['profit']:,.2f}",
                 f"{currency}{summary['prev_profit']:,.2f}",
                 f"{summary['profit_change']:+.1f}%"],
                ["入住次数",
                 str(summary['checkins']),
                 str(summary['prev_checkins']),
                 f"{summary['checkin_change']:+.1f}%"],
                ["平均出租率",
                 f"{summary['avg_occ']:.1f}%",
                 f"{summary['prev_avg_occ']:.1f}%",
                 f"{summary['avg_occ'] - summary['prev_avg_occ']:+.1f}pp"],
            ]
            kpi_table = Table(kpi_data, colWidths=[45*mm, 40*mm, 40*mm, 35*mm])
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(_p("primary"))),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, -1), font_name),
                ("FONTSIZE",    (0, 0), (-1, 0), 10),
                ("FONTSIZE",    (0, 1), (-1, -1), 9),
                ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
                ("ALIGN",       (0, 0), (0, -1), "LEFT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor(_p("bg")), colors.white]),
                ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor(_p("border"))),
                ("TOPPADDING",  (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(kpi_table)
            story.append(Spacer(1, 8*mm))

            # 房型收益表
            if rt["types"]:
                story.append(Paragraph("🏠 房型收益对比", h2_style))
                rt_data = [["房型", "入住次数", f"营收({currency})"]]
                for t, c, r in zip(rt["types"], rt["checkins"], rt["revenues"]):
                    rt_data.append([t, str(c), f"{r:,.2f}"])
                rt_table = Table(rt_data, colWidths=[60*mm, 40*mm, 60*mm])
                rt_table.setStyle(TableStyle([
                    ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(_p("text"))),
                    ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                    ("FONTNAME",    (0, 0), (-1, -1), font_name),
                    ("FONTSIZE",    (0, 0), (-1, -1), 9),
                    ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.HexColor(_p("bg")), colors.white]),
                    ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor(_p("border"))),
                    ("TOPPADDING",  (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(rt_table)
                story.append(Spacer(1, 8*mm))

            # 消费排行榜
            if top:
                story.append(Paragraph("🏆 消费排行 TOP10", h2_style))
                top_data = [["排名", "姓名", "电话", "入住次数", f"累计消费({currency})"]]
                for i, g in enumerate(top, 1):
                    top_data.append([
                        str(i), g["name"], g["phone"],
                        str(g["stays"]), f"{g['total']:,.2f}"
                    ])
                top_table = Table(top_data,
                                  colWidths=[15*mm, 35*mm, 35*mm, 25*mm, 50*mm])
                top_table.setStyle(TableStyle([
                    ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(_p("text"))),
                    ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                    ("FONTNAME",    (0, 0), (-1, -1), font_name),
                    ("FONTSIZE",    (0, 0), (-1, -1), 9),
                    ("ALIGN",       (0, 0), (0, -1), "CENTER"),
                    ("ALIGN",       (3, 0), (-1, -1), "RIGHT"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.HexColor(_p("bg")), colors.white]),
                    ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor(_p("border"))),
                    ("TOPPADDING",  (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    # 第一名金色高亮
                    ("TEXTCOLOR",   (0, 1), (-1, 1), colors.HexColor(_p("accent"))),
                    ("FONTNAME",    (0, 1), (-1, 1), font_name),
                ]))
                story.append(top_table)

            # ── 生成 PDF ──
            doc = SimpleDocTemplate(
                filepath, pagesize=A4,
                leftMargin=20*mm, rightMargin=20*mm,
                topMargin=20*mm, bottomMargin=20*mm
            )
            doc.build(story)
            return True, filepath

        except ImportError:
            # reportlab 未安装，降级为纯文本
            txt_path = filepath.replace(".pdf", "_报表.txt")
            ok, result = ReportExporter.export_monthly_csv(year, month, txt_path)
            if ok:
                return True, f"{txt_path}（reportlab未安装，已降级为文本格式）"
            return False, "reportlab 未安装，且 CSV 降级也失败"
        except Exception as e:
            return False, str(e)


# ─────────────────────────────────────────────
#  报表标签页（嵌入 WorkspaceDock）
# ─────────────────────────────────────────────

class ReportTab(QWidget):
    """
    报表与数据分析标签页
    包含：月度报表 / 年度趋势 / 出租率 / 房型分析
    """

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setObjectName("ReportTab")
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        # ── 标题栏 ──
        title_row = QHBoxLayout()
        title_lbl = QLabel("报表与数据分析")
        self._title_lbl = title_lbl
        title_lbl.setObjectName("H2Title")
        title_row.addWidget(title_lbl)
        title_row.addStretch()

        # 年月选择
        now = datetime.now()
        self.year_combo = QComboBox()
        for y in range(now.year - 2, now.year + 1):
            self.year_combo.addItem(str(y), y)
        self.year_combo.setCurrentIndex(self.year_combo.count() - 1)

        self.month_combo = QComboBox()
        for m in range(1, 13):
            self.month_combo.addItem(f"{m}月", m)
        self.month_combo.setCurrentIndex(now.month - 1)

        title_row.addWidget(QLabel("年份："))
        title_row.addWidget(self.year_combo)
        title_row.addWidget(QLabel("月份："))
        title_row.addWidget(self.month_combo)

        self._btn_refresh = QPushButton("刷新")
        self._btn_refresh.setObjectName("SolidPrimaryBtn")
        self._btn_refresh.clicked.connect(self.refresh)
        title_row.addWidget(self._btn_refresh)

        self._btn_export = QPushButton("导出 Excel")
        self._btn_export.setObjectName("FdActSecondary")
        self._btn_export.clicked.connect(self._export)
        title_row.addWidget(self._btn_export)

        self._btn_pdf = QPushButton("导出 PDF")
        self._btn_pdf.setObjectName("FdDangerBtn")
        self._btn_pdf.clicked.connect(self._export_pdf)
        title_row.addWidget(self._btn_pdf)

        self._btn_integrity = QPushButton("诚信报告 PDF")
        self._btn_integrity.setObjectName("FdActSecondary")
        self._btn_integrity.clicked.connect(self._export_integrity_pdf)
        title_row.addWidget(self._btn_integrity)

        layout.addLayout(title_row)

        # ── KPI 卡片行 ──
        kpi_frame = QFrame()
        kpi_frame.setObjectName("ContentBox")
        from ui_surface import fd_apply_content_box
        fd_apply_content_box(kpi_frame)
        kpi_cl = QVBoxLayout(kpi_frame)
        kpi_cl.setContentsMargins(12, 10, 12, 10)
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(10)
        self.kpi_income  = KpiCard("本月收入", "¥0", "较上月 --", "amount_positive")
        self.kpi_expense = KpiCard("本月支出", "¥0", "较上月 --", "danger")
        self.kpi_profit  = KpiCard("净利润",   "¥0", "较上月 --", "primary")
        self.kpi_occ     = KpiCard("平均出租率", "0%", "近30天", "accent")
        for card in [self.kpi_income, self.kpi_expense, self.kpi_profit, self.kpi_occ]:
            kpi_row.addWidget(card)
        kpi_cl.addLayout(kpi_row)
        layout.addWidget(kpi_frame)

        # ── 图表标签页 ──
        self.chart_tabs = QTabWidget()
        self.chart_tabs.setDocumentMode(True)

        # 月度收支折线图
        self.revenue_chart = LineChartWidget()
        self.chart_tabs.addTab(self.revenue_chart, "月度收支")

        # 年度柱状图
        self.yearly_chart = BarChartWidget()
        self.chart_tabs.addTab(self.yearly_chart, "年度对比")

        # 出租率趋势
        self.occ_chart = LineChartWidget()
        self.chart_tabs.addTab(self.occ_chart, "出租率趋势")

        # 支付方式饼图
        self.pay_chart = PieChartWidget()
        self.chart_tabs.addTab(self.pay_chart, "支付方式")

        # 房型收益柱状图
        self.room_type_chart = BarChartWidget()
        self.chart_tabs.addTab(self.room_type_chart, "房型收益")

        # 消费排行表
        self.top_guest_table = self._build_top_guest_table()
        self.chart_tabs.addTab(self.top_guest_table, "消费排行")

        chart_box = QFrame()
        chart_box.setObjectName("ContentBox")
        chart_l = QVBoxLayout(chart_box)
        chart_l.setContentsMargins(0, 0, 0, 0)
        chart_l.addWidget(self.chart_tabs)
        layout.addWidget(chart_box, 1)

        from ui_surface import fd_apply_content_box, fd_apply_table_palette, fd_connect_theme_refresh
        fd_apply_content_box(chart_box)
        fd_apply_table_palette(self.top_table)
        fd_connect_theme_refresh(self)

    def _refresh_theme_styles(self) -> None:
        for card in (self.kpi_income, self.kpi_expense, self.kpi_profit, self.kpi_occ):
            card.refresh_theme()
        for chart in (
            self.revenue_chart,
            self.yearly_chart,
            self.occ_chart,
            self.pay_chart,
            self.room_type_chart,
        ):
            chart.update()

    def _build_top_guest_table(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(0, 8, 0, 0)

        lbl = QLabel("消费金额 TOP 10 住客")
        self._top_guest_lbl = lbl
        lbl.setObjectName("H3Title")
        layout.addWidget(lbl)

        self.top_table = QTableWidget()
        self.top_table.setColumnCount(4)
        self.top_table.setHorizontalHeaderLabels(["姓名", "电话", "入住次数", "累计消费"])
        self.top_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.top_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.top_table.setAlternatingRowColors(False)
        self.top_table.verticalHeader().setVisible(False)
        layout.addWidget(self.top_table)
        return w

    def refresh(self) -> None:
        """刷新所有报表数据"""
        try:
            from vendor_lockdown import is_locked, lock_message
            if is_locked("reports"):
                msg = lock_message()
                self.kpi_income.update_value("已锁定", msg[:20])
                self.kpi_expense.update_value("已锁定", "")
                self.kpi_profit.update_value("已锁定", "")
                self.kpi_occ.update_value("已锁定", "")
                return
        except Exception:
            pass
        year  = self.year_combo.currentData()
        month = self.month_combo.currentData()
        currency = db.get_config("currency") or "¥"

        try:
            # ── KPI ──
            data = ReportData.monthly_revenue(year, month)
            self.kpi_income.update_value(f"{currency}{data['total_income']:,.0f}")
            self.kpi_expense.update_value(f"{currency}{data['total_expense']:,.0f}")
            net = data["net_profit"]
            self.kpi_profit.update_value(
                f"{currency}{net:,.0f}",
                "盈利 ✅" if net >= 0 else "亏损 ⚠️"
            )

            occ = ReportData.occupancy_trend(30)
            self.kpi_occ.update_value(f"{occ['avg_rate']}%", f"共{occ['total_rooms']}间房")

            # ── 月度收支折线图 ──
            if data["days"]:
                self.revenue_chart.set_data(
                    f"{year}年{month}月 每日收支",
                    [d[5:] for d in data["days"]],  # 只显示 MM-DD
                    [
                        {"name": "收入", "data": data["incomes"],
                         "color": QColor(_p("amount_positive"))},
                        {"name": "支出", "data": data["expenses"],
                         "color": QColor(_p("danger"))},
                    ],
                    y_unit=currency
                )

            # ── 年度柱状图 ──
            yearly = ReportData.yearly_revenue(year)
            self.yearly_chart.set_data(
                f"{year}年 月度营收对比",
                yearly["months"],
                [
                    {"name": "收入", "data": yearly["incomes"],
                     "color": QColor(_p("primary"))},
                    {"name": "净利润", "data": yearly["net_profits"],
                     "color": QColor(_p("amount_positive"))},
                ],
                y_unit=currency
            )

            # ── 出租率趋势 ──
            self.occ_chart.set_data(
                "近30天出租率趋势",
                occ["days"],
                [{"name": "出租率", "data": occ["rates"],
                  "color": QColor(_p("warn"))}],
                y_unit="%"
            )

            # ── 支付方式饼图 ──
            pay = ReportData.payment_method_breakdown(year, month)
            if pay["methods"]:
                self.pay_chart.set_data(
                    f"{year}年{month}月 支付方式分布",
                    pay["methods"],
                    pay["amounts"]
                )

            # ── 房型收益 ──
            rt = ReportData.room_type_revenue(year, month)
            if rt["types"]:
                self.room_type_chart.set_data(
                    f"{year}年{month}月 房型收益对比",
                    rt["types"],
                    [{"name": "营收", "data": rt["revenues"],
                      "color": QColor(_p("primary"))}],
                    y_unit=currency
                )

            # ── 消费排行 ──
            top = ReportData.top_guests(10)
            self.top_table.setRowCount(len(top))
            for r, g in enumerate(top):
                self.top_table.setItem(r, 0, QTableWidgetItem(g["name"]))
                self.top_table.setItem(r, 1, QTableWidgetItem(g["phone"]))
                self.top_table.setItem(r, 2, QTableWidgetItem(str(g["stays"])))
                item = QTableWidgetItem(f"{currency}{g['total']:,.2f}")
                if r == 0:
                    item.setForeground(QColor(_p("accent")))
                    item.setFont(QFont("", -1, QFont.Bold))
                self.top_table.setItem(r, 3, item)

        except Exception as e:
            pass  # 静默失败

    def _export(self) -> None:
        """导出报表"""
        year  = self.year_combo.currentData()
        month = self.month_combo.currentData()

        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "导出报表",
            f"报表_{year}年{month}月.xlsx",
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv)"
        )
        if not filepath:
            return

        if filepath.endswith(".csv"):
            ok, result = ReportExporter.export_monthly_csv(year, month, filepath)
        else:
            ok, result = ReportExporter.export_monthly_excel(year, month, filepath)

        if ok:
            show_info(self, f"✅ 报表已导出：\n{result}")
        else:
            show_error(self, "导出失败", result)

    def _export_pdf(self) -> None:
        """导出 PDF 报表"""
        year  = self.year_combo.currentData()
        month = self.month_combo.currentData()

        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "导出 PDF 报表",
            f"报表_{year}年{month}月.pdf",
            "PDF 文件 (*.pdf)"
        )
        if not filepath:
            return

        ok, result = ReportExporter.export_pdf(year, month, filepath)
        if ok:
            show_info(self, f"✅ PDF 报表已导出：\n{result}")
        else:
            show_error(self, "PDF 导出失败", result)

    def _export_integrity_pdf(self) -> None:
        """导出月度诚信 PDF。"""
        year = self.year_combo.currentData()
        month = self.month_combo.currentData()
        filepath, _ = QFileDialog.getSaveFileName(
            self,
            "导出月度诚信报告",
            f"诚信报告_{year}年{month}月.pdf",
            "PDF 文件 (*.pdf)"
        )
        if not filepath:
            return
        from integrity_report import export_integrity_pdf
        ok, result = export_integrity_pdf(year, month, filepath)
        if ok:
            show_info(self, f"✅ 诚信报告已导出：\n{result}")
        else:
            show_error(self, "诚信报告导出失败", result)
