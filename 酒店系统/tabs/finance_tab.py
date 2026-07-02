from __future__ import annotations

import datetime
from collections import defaultdict

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox,
    QLineEdit, QFormLayout, QTableWidget, QTableWidgetItem, QHeaderView,
    QGridLayout, QFrame, QScrollArea, QSplitter, QDialog,
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor

from database import db, LEDGER_REVENUE_TX_TYPES, LEDGER_DEPOSIT_TX_TYPES
from event_bus import bus
from i18n import i18n
from ui_helpers import ask_confirm, show_info, show_warning, style_dialog, build_dialog_header
from design_tokens import _p
from frontdesk_ui import (
    fd_section_bar, FD_MARGIN, fd_apply_low_freq_btn, fd_apply_card_action_btn,
    fd_apply_action_btn, FD_SPACE_SM,
)
from ui_surface import fd_apply_card_panel, fd_apply_content_box, fd_apply_table_palette, fd_refresh_surfaces, fd_apply_panel_sep, fd_apply_scroll_area, fd_sync_table_height, fd_apply_workspace_splitter, fd_apply_page_tab_root
from ledger_format import ledger_tx_type_display


# ═══════════════════════════════════════════════════════════
# 支出管理与盈亏标签页（第 3 版 — 设计系统 v3）
# ═══════════════════════════════════════════════════════════
# 支出类型：账本备注前缀使用英文编码（如 UTIL:备注），便于分类且与界面语言解耦
EXPENSE_TYPES = [
    ("UTIL", "expense_util"),
    ("REPAIR", "expense_repair"),
    ("PURCHASE", "expense_purchase"),
    ("LINEN_PURCHASE", "expense_linen_purchase"),
    ("SALARY", "expense_salary"),
    ("WITHDRAW", "expense_withdraw"),
    ("MISC", "expense_misc"),
]
_EXPENSE_CODE_TO_KEY = dict(EXPENSE_TYPES)

PAYMENT_METHODS = [
    ("CASH_USD", "", "payment_usd_cash", "USD"),
    ("CASH_KHR", "៛", "payment_khr_cash", "KHR"),
    ("USDT", "", "payment_usdt", "TRC20"),
    ("ABA", "", "payment_aba", "QR / Card"),
]
PAYMENT_LEGACY_LABELS = {
    "CASH": "legacy_cash",
    "CARD": "legacy_card",
    "TRANSFER": "legacy_transfer",
    "WECHAT": "legacy_wechat",
    "ALIPAY": "legacy_alipay",
    "SYSTEM": "payment_legacy_system",
}


def pay_method_label(code: str | None) -> str:
    c = (code or "CASH_USD").strip()
    for method_code, _icon, label_key, _sub in PAYMENT_METHODS:
        if c == method_code:
            return i18n.t(label_key)
    return i18n.t(PAYMENT_LEGACY_LABELS.get(c, c)) if PAYMENT_LEGACY_LABELS.get(c) else (c or "-")


def normalize_pay_method(code: str | None) -> str:
    c = (code or "CASH_USD").strip()
    if c in {"CASH", "TRANSFER"}:
        return "CASH_USD"
    if c in {"CARD", "WECHAT", "ALIPAY"}:
        return "ABA"
    return c


def _finance_pay_methods_combo():
    cmb = QComboBox()
    for code, _icon, label_key, sub in PAYMENT_METHODS:
        cmb.addItem(f"{i18n.t(label_key)} · {sub}", code)
    return cmb


def _checkin_pay_methods_combo():
    cmb = QComboBox()
    for code, _icon, label_key, sub in PAYMENT_METHODS:
        cmb.addItem(f"{i18n.t(label_key)} · {sub}", code)
    return cmb


def _payout_category_from_note(note):
    """Resolve ledger note to expense CODE or OTHER (supports CODE:detail and legacy Chinese labels)."""
    if not note:
        return "OTHER"
    head = str(note).split(":", 1)[0].strip()
    if head in _EXPENSE_CODE_TO_KEY:
        return head
    legacy = (
        ("UTIL", ("水电", "")),
        ("REPAIR", ("维修", "")),
        ("PURCHASE", ("物料", "采购", "")),
        ("LINEN_PURCHASE", ("布草", "洗衣", "linen", "")),
        ("SALARY", ("工资", "")),
        ("WITHDRAW", ("取款", "老板", "")),
        ("MISC", ("杂项", "")),
    )
    for code, needles in legacy:
        if any(n in head for n in needles):
            return code
    return "OTHER"


class FinanceTab(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName("FinanceTab")
        self._range_key = "today"
        l = QVBoxLayout(self)
        l.setSpacing(FD_SPACE_SM)
        l.setContentsMargins(FD_MARGIN, FD_MARGIN, FD_MARGIN, FD_MARGIN)

        # ── 顶部金线横栏：标题 + 时间范围 + 操作按钮 ──
        btn_rf = QPushButton(i18n.t("btn_reload"))
        fd_apply_low_freq_btn(btn_rf)
        btn_rf.clicked.connect(self.refresh)

        btn_export = QPushButton(i18n.t("btn_export_ledger"))
        fd_apply_card_action_btn(btn_export)
        btn_export.clicked.connect(self._export_ledger)

        btn_payout = QPushButton(i18n.t("finance_btn_register_expense"))
        fd_apply_action_btn(btn_payout, primary=True)
        btn_payout.clicked.connect(self._do_payout)

        # [sub-d Task1] 历史账单查询入口：弹 BillDetailDialog 搜索模式
        btn_bill_query = QPushButton("历史账单")
        fd_apply_low_freq_btn(btn_bill_query)
        btn_bill_query.clicked.connect(self._open_bill_history)

        # [sub-d Task3] 多币种对账 / 多币种月报入口
        btn_multi_reconcile = QPushButton("多币种对账")
        fd_apply_low_freq_btn(btn_multi_reconcile)
        btn_multi_reconcile.clicked.connect(self._open_multi_reconcile)

        btn_multi_report = QPushButton("多币种月报")
        fd_apply_low_freq_btn(btn_multi_report)
        btn_multi_report.clicked.connect(self._open_multi_report)

        self.cmb_range = QComboBox()
        self.cmb_range.setObjectName("FdCompactCombo")
        for label_key, key in (
            ("range_today", "today"),
            ("range_yesterday", "yesterday"),
            ("range_week", "week"),
            ("range_month", "month"),
            ("range_last_month", "last_month"),
        ):
            self.cmb_range.addItem(i18n.t(label_key), key)
        self.cmb_range.currentIndexChanged.connect(self.refresh)

        l.addWidget(fd_section_bar(
            i18n.t("finance_title_bar"),
            action_widgets=[
                self.cmb_range, btn_rf, btn_export,
                btn_bill_query, btn_multi_reconcile, btn_multi_report,
                btn_payout,
            ],
        ))

        # ── 统计卡片 — 分 3 组 + VLine 隔开 ──
        self._cards = {}
        group1_defs = [
            ("today_income", "finance_card_today_in", "FdAmountPositive"),
            ("today_expense", "finance_card_today_out", "FdAmountWarn"),
            ("today_profit", "finance_card_today_net", "FdAmountPrimary"),
        ]
        group2_defs = [
            ("month_income", "finance_card_month_in", "FdAmountPositive"),
            ("month_expense", "finance_card_month_out", "FdAmountWarn"),
        ]
        group3_defs = [
            ("deposit_pool", "finance_card_deposit", "FdAmountPrimary"),
            ("fund_pool", "finance_card_fund", "FdAmountPrimary"),
        ]

        def _build_kpi_row(defs, group: str = "secondary"):
            # G03: 每组 KPI 带 kpi_group 动态属性，QSS 可按组设不同字号/权重
            row = QHBoxLayout()
            row.setSpacing(FD_SPACE_SM)
            for key, lbl_key, amount_style in defs:
                f = QFrame()
                f.setObjectName("FinanceStatCell")
                f.setProperty("kpi_group", group)  # G03: primary/secondary/tertiary
                fl = QVBoxLayout(f)
                fl.setContentsMargins(0, 0, 0, 0)
                fl.setSpacing(2)
                cap = QLabel(i18n.t(lbl_key))
                cap.setObjectName("FdMutedLabel")
                fl.addWidget(cap)
                v = QLabel(i18n.t("currency_symbol") + "0.00")
                v.setObjectName(amount_style)
                self._cards[key] = v
                fl.addWidget(v)
                row.addWidget(f, 1)
            return row

        from ui_surface import fd_apply_panel_sep
        kpi_line = QHBoxLayout()
        kpi_line.setSpacing(0)
        kpi_line.addLayout(_build_kpi_row(group1_defs, "primary"), 2)
        sep1 = QFrame()
        sep1.setObjectName("FdPanelSep")
        fd_apply_panel_sep(sep1)
        kpi_line.addWidget(sep1)
        kpi_line.addLayout(_build_kpi_row(group2_defs, "secondary"), 1)
        sep2 = QFrame()
        sep2.setObjectName("FdPanelSep")
        fd_apply_panel_sep(sep2)
        kpi_line.addWidget(sep2)
        kpi_line.addLayout(_build_kpi_row(group3_defs, "tertiary"), 1)
        l.addLayout(kpi_line)

        # ── LEFT-RIGHT QSplitter：左 = 流水（钱箱），右 = P&L + 对账 + 支出 ──
        split = QSplitter(Qt.Orientation.Horizontal)
        split.setChildrenCollapsible(False)
        split.setHandleWidth(2)
        split.setObjectName("FinanceMainSplitter")

        # ── 左侧面板：流水列表 ──
        left_panel = QFrame()
        left_panel.setObjectName("ContentBox")
        fd_apply_content_box(left_panel)
        left_lay = QVBoxLayout(left_panel)
        left_lay.setContentsMargins(10, 10, 10, 10)
        left_lay.setSpacing(FD_SPACE_SM)

        left_lay.addWidget(fd_section_bar(i18n.t("finance_group_ledger_box")))
        filter_row = QHBoxLayout()
        filter_row.setSpacing(FD_SPACE_SM)
        self.cmb_ledger_type = QComboBox()
        self.cmb_ledger_type.setObjectName("FdCompactCombo")
        for label_key, data in (
            ("finance_filter_all", "ALL"),
            ("finance_filter_revenue", "REVENUE"),
            ("finance_filter_deposit", "DEPOSIT"),
            ("finance_filter_payout", "PAYOUT"),
            ("finance_filter_reconcile", "RECONCILE"),
            ("finance_filter_util", "EXP:UTIL"),
            ("finance_filter_repair", "EXP:REPAIR"),
            ("finance_filter_purchase", "EXP:PURCHASE"),
            ("finance_filter_linen", "EXP:LINEN_PURCHASE"),
            ("finance_filter_salary", "EXP:SALARY"),
            ("finance_filter_withdraw", "EXP:WITHDRAW"),
            ("finance_filter_misc", "EXP:MISC"),
        ):
            self.cmb_ledger_type.addItem(i18n.t(label_key), data)
        self.cmb_ledger_type.currentIndexChanged.connect(self.refresh)
        filter_row.addWidget(self.cmb_ledger_type)
        self.cmb_ledger_pay = QComboBox()
        self.cmb_ledger_pay.setObjectName("FdCompactCombo")
        self.cmb_ledger_pay.addItem(i18n.t("finance_filter_all_payment"), "ALL")
        for code, _icon, label_key, sub in PAYMENT_METHODS:
            self.cmb_ledger_pay.addItem(f"{i18n.t(label_key)} · {sub}", code)
        for code, label_key in PAYMENT_LEGACY_LABELS.items():
            self.cmb_ledger_pay.addItem(i18n.t(label_key), code)
        self.cmb_ledger_pay.currentIndexChanged.connect(self.refresh)
        filter_row.addWidget(self.cmb_ledger_pay)
        self.txt_ledger_search = QLineEdit()
        self.txt_ledger_search.setObjectName("FdCompactInput")
        self.txt_ledger_search.setPlaceholderText(i18n.t("search_ledger_ph"))
        self.txt_ledger_search.setClearButtonEnabled(True)
        self._search_debounce = QTimer(self)
        self._search_debounce.setSingleShot(True)
        self._search_debounce.setInterval(200)
        self._search_debounce.timeout.connect(self.refresh)
        self.txt_ledger_search.textChanged.connect(lambda _t: self._search_debounce.start())
        filter_row.addWidget(self.txt_ledger_search, 1)
        left_lay.addLayout(filter_row)
        self.tbl_ledger = QTableWidget(0, 6)
        self.tbl_ledger.setObjectName("FinanceLedgerTable")
        self.tbl_ledger.setHorizontalHeaderLabels(
            [
                i18n.t("finance_ledger_col_time"),
                i18n.t("finance_ledger_col_type"),
                i18n.t("finance_ledger_col_method"),
                i18n.t("finance_ledger_col_room"),
                i18n.t("finance_ledger_col_amount"),
                i18n.t("finance_ledger_col_note"),
            ]
        )
        led_hdr = self.tbl_ledger.horizontalHeader()
        led_hdr.setMinimumSectionSize(70)
        for c in (0, 1, 2, 3):
            led_hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        led_hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        led_hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.tbl_ledger.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl_ledger.setAlternatingRowColors(False)
        self.tbl_ledger.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        left_lay.addWidget(self.tbl_ledger)
        split.addWidget(left_panel)

        # ── 右侧面板：P&L 摘要 + 对账 + 支出分类（滚动）──
        right_panel = QWidget()
        right_lay = QVBoxLayout(right_panel)
        right_lay.setContentsMargins(0, 0, 0, 0)
        right_lay.setSpacing(0)

        right_scroll = QScrollArea()
        right_scroll.setObjectName("FinanceRightScroll")
        right_scroll.setWidgetResizable(True)
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        right_box = QFrame()
        right_box.setObjectName("ContentBox")
        fd_apply_content_box(right_box)
        right_scroll_lay = QVBoxLayout(right_box)
        right_scroll_lay.setContentsMargins(10, 10, 10, 10)
        right_scroll_lay.setSpacing(FD_SPACE_SM)

        # 操作员统计
        right_scroll_lay.addWidget(fd_section_bar(i18n.t("section_operator_summary")))
        self.tbl_operator_summary = QTableWidget(0, 6)
        self.tbl_operator_summary.setObjectName("FinanceOpSummaryTable")
        self.tbl_operator_summary.setHorizontalHeaderLabels(
            [i18n.t("finance_op_col_operator"), i18n.t("finance_op_col_ops"),
             i18n.t("finance_op_col_issue"), i18n.t("finance_op_col_cancel"),
             i18n.t("finance_op_col_ci"), i18n.t("finance_op_col_last")]
        )
        ops_hdr = self.tbl_operator_summary.horizontalHeader()
        ops_hdr.setMinimumSectionSize(70)
        ops_hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in range(1, 6):
            ops_hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl_operator_summary.verticalHeader().setVisible(False)
        self.tbl_operator_summary.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl_operator_summary.setAlternatingRowColors(False)
        self.tbl_operator_summary.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tbl_operator_summary.setSortingEnabled(True)
        self.tbl_operator_summary.itemDoubleClicked.connect(self._show_operator_detail)
        right_scroll_lay.addWidget(self.tbl_operator_summary)

        # 对账
        right_scroll_lay.addWidget(self._build_reconcile_group())

        # 快捷支出（折叠在 QGroupBox 中）
        quick_exp_group = QFrame()
        quick_exp_group.setObjectName("FdCard")
        quick_exp_lay = QVBoxLayout(quick_exp_group)
        quick_exp_lay.setContentsMargins(12, 8, 12, 8)
        quick_exp_lay.setSpacing(6)
        quick_exp_lay.addWidget(fd_section_bar(i18n.t("finance_quick_group_title")))
        qg = QGridLayout()
        qg.setSpacing(6)
        expense_icons = {
            "UTIL": "",
            "REPAIR": "",
            "PURCHASE": "",
            "LINEN_PURCHASE": "",
            "SALARY": "",
            "WITHDRAW": "",
            "MISC": "",
        }
        for idx, (code, msg_key) in enumerate(EXPENSE_TYPES):
            btn = QPushButton(i18n.t(msg_key))
            btn.setObjectName("FinanceQuickExpenseBtn")
            btn.setFixedHeight(32)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked=False, c=code, mk=msg_key: self._quick_expense(c, mk))
            qg.addWidget(btn, idx // 4, idx % 4)
        quick_exp_lay.addLayout(qg)
        right_scroll_lay.addWidget(quick_exp_group)

        # 支出分类
        right_scroll_lay.addWidget(fd_section_bar(i18n.t("finance_cat_today_title")))
        self.tbl_cat = QTableWidget(0, 3)
        self.tbl_cat.setObjectName("FinanceCatTable")
        self.tbl_cat.setHorizontalHeaderLabels(
            [
                i18n.t("finance_cat_col_type"),
                i18n.t("finance_cat_col_count"),
                i18n.t("finance_cat_col_amt"),
            ]
        )
        cat_hdr = self.tbl_cat.horizontalHeader()
        cat_hdr.setMinimumSectionSize(70)
        cat_hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        cat_hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        cat_hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.tbl_cat.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl_cat.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tbl_cat.itemDoubleClicked.connect(self._apply_category_from_table)
        right_scroll_lay.addWidget(self.tbl_cat)

        right_scroll.setWidget(right_box)
        fd_apply_scroll_area(right_scroll)
        right_lay.addWidget(right_scroll, 1)
        split.addWidget(right_panel)
        # [F10] 左（流水列表）:右（报表+对账）比例从 3:2 改为 7:5
        # 左侧 6 列流水表（时间/类型/方式/房间/金额/备注）需要更多宽度
        split.setStretchFactor(0, 7)
        split.setStretchFactor(1, 5)
        split.setSizes([700, 500])
        l.addWidget(split, stretch=1)

        for _tbl in (
            self.tbl_ledger, self.tbl_operator_summary,
            self.tbl_cat, self.tbl_reconcile,
        ):
            fd_apply_table_palette(_tbl)
        fd_apply_card_panel(quick_exp_group)
        fd_apply_workspace_splitter(split)
        fd_apply_page_tab_root(self)

        fd_refresh_surfaces(self)
        self.refresh()
        bus.theme_changed.connect(lambda _: fd_refresh_surfaces(self))

    def _period_where(self, column: str = "created_at") -> str:
        key = self.cmb_range.currentData() if hasattr(self, "cmb_range") else "today"
        if key == "yesterday":
            return f"date({column})=date('now','localtime','-1 day')"
        if key == "week":
            return f"date({column})>=date('now','localtime','weekday 0','-6 days')"
        if key == "month":
            return f"strftime('%Y-%m', {column})=strftime('%Y-%m', 'now', 'localtime')"
        if key == "last_month":
            return f"strftime('%Y-%m', {column})=strftime('%Y-%m', 'now', 'localtime','start of month','-1 month')"
        return f"date({column})=date('now','localtime')"

    def _build_reconcile_group(self) -> QFrame:
        gb = QFrame()
        gb.setObjectName("FdReconcileGroup")
        lay = QVBoxLayout(gb)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(FD_SPACE_SM)
        lay.addWidget(fd_section_bar(i18n.t("section_reconcile_today")))
        self.tbl_reconcile = QTableWidget(0, 4)
        self.tbl_reconcile.setObjectName("FinanceReconcileTable")
        self.tbl_reconcile.setHorizontalHeaderLabels(
            [i18n.t("reconcile_col_method"), i18n.t("reconcile_col_expected"),
             i18n.t("reconcile_col_actual"), i18n.t("reconcile_col_diff")]
        )
        rec_hdr = self.tbl_reconcile.horizontalHeader()
        rec_hdr.setMinimumSectionSize(70)
        rec_hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for c in (1, 2, 3):
            rec_hdr.setSectionResizeMode(c, QHeaderView.ResizeMode.ResizeToContents)
        self.tbl_reconcile.verticalHeader().setVisible(False)
        self.tbl_reconcile.setEditTriggers(
            QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.EditKeyPressed
        )
        self.tbl_reconcile.itemChanged.connect(self._on_reconcile_cell_changed)
        lay.addWidget(self.tbl_reconcile)
        self.btn_reconcile_done = QPushButton(i18n.t("btn_finish_reconcile"))
        fd_apply_card_action_btn(self.btn_reconcile_done)
        self.btn_reconcile_done.clicked.connect(self._finish_reconcile)
        lay.addWidget(self.btn_reconcile_done)
        return gb

    def _populate_reconcile(self, expected: dict[str, float]) -> None:
        cur = i18n.t("currency_symbol")
        self.tbl_reconcile.blockSignals(True)
        self.tbl_reconcile.setRowCount(0)
        try:
            for row, (code, icon, label_key, sub) in enumerate(PAYMENT_METHODS):
                amount = float(expected.get(code, 0.0) or 0.0)
                self.tbl_reconcile.insertRow(row)
                self.tbl_reconcile.setItem(row, 0, QTableWidgetItem(f"{icon} {i18n.t(label_key)}"))
                exp_item = QTableWidgetItem(f"{cur}{amount:.2f}")
                exp_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.tbl_reconcile.setItem(row, 1, exp_item)
                act_item = QTableWidgetItem(f"{cur}{amount:.2f}")
                act_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                act_item.setData(Qt.ItemDataRole.UserRole, code)
                act_item.setFlags(
                    act_item.flags()
                    | Qt.ItemFlag.ItemIsEditable
                    | Qt.ItemFlag.ItemIsEnabled
                    | Qt.ItemFlag.ItemIsSelectable
                )
                self.tbl_reconcile.setItem(row, 2, act_item)
                diff_item = QTableWidgetItem("")
                diff_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                diff_item.setData(Qt.ItemDataRole.UserRole, code)
                self.tbl_reconcile.setItem(row, 3, diff_item)
                self._refresh_reconcile_diff(code)
        finally:
            self.tbl_reconcile.blockSignals(False)
        fd_sync_table_height(
            self.tbl_reconcile,
            min_rows=len(PAYMENT_METHODS),
            max_rows=len(PAYMENT_METHODS),
        )

    def _on_reconcile_cell_changed(self, item: QTableWidgetItem) -> None:
        if item is None or item.column() != 2:
            return
        code = item.data(Qt.ItemDataRole.UserRole)
        if not code:
            return
        self._refresh_reconcile_diff(str(code))

    def _refresh_reconcile_diff(self, code: str) -> None:
        cur = i18n.t("currency_symbol")
        row = next((r for r in range(self.tbl_reconcile.rowCount()) if self.tbl_reconcile.item(r, 3).data(Qt.ItemDataRole.UserRole) == code), -1)
        if row < 0:
            return
        expected_txt = self.tbl_reconcile.item(row, 1).text().replace(cur, "").replace(",", "")
        try:
            expected = float(expected_txt)
        except Exception:
            expected = 0.0
        raw = self.tbl_reconcile.item(row, 2).text().replace(cur, "").replace(",", "").strip()
        try:
            actual = float(raw)
        except Exception:
            actual = 0.0
        diff = actual - expected
        item = self.tbl_reconcile.item(row, 3)
        item.setText(f"{cur}{diff:.2f}")
        item.setForeground(QColor(_p("amount_positive") if diff >= 0 else _p("danger")))

    def _finish_reconcile(self) -> None:
        lines = []
        for row in range(self.tbl_reconcile.rowCount()):
            method = self.tbl_reconcile.item(row, 0).text()
            expected = self.tbl_reconcile.item(row, 1).text()
            actual = self.tbl_reconcile.item(row, 2).text()
            diff = self.tbl_reconcile.item(row, 3).text()
            lines.append(i18n.t("reconcile_record_line").format(method=method, expected=expected, actual=actual, diff=diff))
        note = ";".join(lines)
        try:
            db.append_ledger("CASH_RECONCILE", 0, "SYSTEM", 1, note=note, pay_method="SYSTEM")
        except Exception:
            show_warning(self, i18n.t("dlg_error"), i18n.t("reconcile_save_failed"))
        bus.show_success_overlay.emit(i18n.t("msg_reconcile_recorded"))
        self.refresh()

    def _apply_category_from_table(self, item):
        if not item:
            return
        code = item.data(Qt.ItemDataRole.UserRole)
        if not code:
            return
        idx = self.cmb_ledger_type.findData(f"EXP:{code}")
        if idx >= 0:
            self.cmb_ledger_type.setCurrentIndex(idx)

    def _export_ledger(self):
        try:
            from openpyxl import Workbook
            from pathlib import Path
            import os
            wb = Workbook()
            ws = wb.active
            ws.title = "ledger"
            headers = [self.tbl_ledger.horizontalHeaderItem(i).text() for i in range(self.tbl_ledger.columnCount())]
            ws.append(headers)
            for r in range(self.tbl_ledger.rowCount()):
                ws.append([
                    self.tbl_ledger.item(r, c).text() if self.tbl_ledger.item(r, c) else ""
                    for c in range(self.tbl_ledger.columnCount())
                ])
            out_dir = Path(db.get_config("report_export_dir") or "reports")
            out_dir.mkdir(exist_ok=True)
            path = out_dir / f"finance_ledger_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(path)
            show_info(self, i18n.t("dlg_export_done"), i18n.t("dlg_export_path").format(path=os.path.abspath(path)))
        except Exception as exc:
            show_warning(self, i18n.t("dlg_export_failed"), str(exc))

    # ── 快捷支出：弹出金额+备注对话框 ──
    def _quick_expense(self, code: str, msg_key: str):
        label = i18n.t(msg_key)
        d = QDialog(self)
        d.setWindowTitle(i18n.t("finance_quick_entry_title").format(label))
        style_dialog(d, size="compact")
        lv = QVBoxLayout(d); lv.setContentsMargins(FD_MARGIN, FD_MARGIN, FD_MARGIN, FD_MARGIN); lv.setSpacing(FD_SPACE_SM)
        lv.addWidget(build_dialog_header(label, i18n.t("finance_quick_entry_subtitle")))
        f = QFormLayout()
        spn_amt = QDoubleSpinBox()
        spn_amt.setRange(0.01, 999999); spn_amt.setDecimals(2)
        spn_amt.setPrefix(i18n.t("currency_symbol")); spn_amt.setValue(0.01)
        cmb_pay = _finance_pay_methods_combo()
        txt_note = QLineEdit(); txt_note.setPlaceholderText(i18n.t("finance_note_optional_ph"))
        f.addRow(i18n.t("finance_field_amount"), spn_amt)
        f.addRow(i18n.t("finance_field_pay_method"), cmb_pay)
        f.addRow(i18n.t("finance_field_note_detail"), txt_note)
        btn_ok = QPushButton(i18n.t("finance_confirm_register")); btn_ok.setObjectName("SolidPrimaryBtn"); btn_ok.clicked.connect(d.accept)
        f.addRow(btn_ok)
        lv.addLayout(f)
        if not d.exec():
            return
        amt = spn_amt.value()
        pay_method = cmb_pay.currentData() or "CASH_USD"
        cur = i18n.t("currency_symbol")
        if amt >= 300 and not ask_confirm(self, i18n.t("dlg_large_expense_confirm"), i18n.t("msg_large_expense").format(cur=cur, amt=amt, label=label, pay=pay_method_label(pay_method))):
            return
        note_text = txt_note.text().strip()
        full_note = f"{code}:{note_text}" if note_text else code
        try:
            db.append_ledger("PAYOUT", -amt, "CASH", 1, note=full_note, pay_method=pay_method)
            bus.show_success_overlay.emit(i18n.t("finance_expense_registered").format(cur, amt))
            self.refresh()
            try:
                from telegram_shadow import telegram_thread
                if telegram_thread.isRunning():
                    telegram_thread.notify_payout(label, amt, note_text)
            except Exception:
                pass
        except Exception as e:
            show_warning(self, i18n.t("finance_register_failed"), str(e))

    def refresh(self):
        cur = i18n.t("currency_symbol")
        today_filter = self._period_where()
        month_filter = "strftime('%Y-%m', created_at)=strftime('%Y-%m', 'now', 'localtime')"
        rev_in = ",".join(f"'{t}'" for t in LEDGER_REVENUE_TX_TYPES)
        dep_in = ",".join(f"'{t}'" for t in LEDGER_DEPOSIT_TX_TYPES)

        # 今日营业额（不含押金）
        try:
            income_today = float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type IN ({rev_in}) AND {today_filter}"
            ).fetchone()[0] or 0)
        except Exception:
            income_today = 0.0
        # 今日支出
        try:
            expense_today = abs(float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type='PAYOUT' AND {today_filter}"
            ).fetchone()[0] or 0))
        except Exception:
            expense_today = 0.0
        # 本月营业额
        try:
            income_month = float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type IN ({rev_in}) AND {month_filter}"
            ).fetchone()[0] or 0)
        except Exception:
            income_month = 0.0
        # 本月支出
        try:
            expense_month = abs(float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type='PAYOUT' AND {month_filter}"
            ).fetchone()[0] or 0))
        except Exception:
            expense_month = 0.0
        try:
            dep_today = float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type IN ({dep_in}) AND {today_filter}"
            ).fetchone()[0] or 0)
        except Exception:
            dep_today = 0.0
        try:
            dep_month = float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type IN ({dep_in}) AND {month_filter}"
            ).fetchone()[0] or 0)
        except Exception:
            dep_month = 0.0

        profit_today = income_today - expense_today
        try:
            deposit_pool = float(db.execute(
                f"SELECT COALESCE(SUM(amount),0) FROM ledger WHERE tx_type IN ({dep_in})"
            ).fetchone()[0] or 0)
        except Exception:
            deposit_pool = 0.0
        try:
            fund = db.get_fund_pool()
        except Exception:
            fund = 0.0

        # G04: KPI 数字变化动效 (#3 节奏: ease-out 120ms 淡出旧值 → 新值)
        from motion_gate import animate_kpi
        animate_kpi(self._cards["today_income"],  f"{cur}{income_today:.2f}")
        animate_kpi(self._cards["today_expense"],  f"{cur}{expense_today:.2f}")
        animate_kpi(self._cards["today_profit"],   f"{cur}{profit_today:.2f}")
        animate_kpi(self._cards["month_income"],   f"{cur}{income_month:.2f}")
        animate_kpi(self._cards["month_expense"],  f"{cur}{expense_month:.2f}")
        animate_kpi(self._cards["deposit_pool"],   f"{cur}{deposit_pool:.2f}")
        animate_kpi(self._cards["fund_pool"],      f"{cur}{fund:.2f}")

        # 当前时间范围支出分类（备注前缀编码:…，兼容旧版中文标签）
        self.tbl_cat.setRowCount(0)
        try:
            agg = defaultdict(lambda: [0, 0.0])
            for note, amt in db.execute(
                f"SELECT note, amount FROM ledger WHERE tx_type='PAYOUT' AND {today_filter}"
            ).fetchall():
                cat = _payout_category_from_note(note)
                agg[cat][0] += 1
                agg[cat][1] += abs(float(amt or 0))
            rows_sorted = sorted(agg.items(), key=lambda x: -x[1][1])
            for i, (cat, (cnt, amt)) in enumerate(rows_sorted):
                display = i18n.t("expense_other") if cat == "OTHER" else i18n.t(_EXPENSE_CODE_TO_KEY[cat])
                self.tbl_cat.insertRow(i)
                cat_item = QTableWidgetItem(display[:40])
                cat_item.setData(Qt.ItemDataRole.UserRole, cat)
                self.tbl_cat.setItem(i, 0, cat_item)
                self.tbl_cat.setItem(i, 1, QTableWidgetItem(str(cnt)))
                self.tbl_cat.setItem(i, 2, QTableWidgetItem(f"{cur}{amt:.2f}"))
                cat_amt_item = self.tbl_cat.item(i, 2)
                if cat_amt_item:
                    cat_amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        except Exception:
            pass

        # 按支付方式对账（营业额 + 押金，不含支出）
        expected_by_method = defaultdict(float)
        try:
            for pay, amt in db.execute(
                f"SELECT COALESCE(pay_method,'CASH_USD'), COALESCE(SUM(amount),0) "
                f"FROM ledger WHERE tx_type IN ({rev_in},{dep_in}) AND {today_filter} GROUP BY pay_method"
            ).fetchall():
                expected_by_method[normalize_pay_method(pay)] += float(amt or 0)
        except Exception:
            pass
        if hasattr(self, "tbl_reconcile"):
            self._populate_reconcile(expected_by_method)
        if hasattr(self, "tbl_operator_summary"):
            self._populate_operator_summary()

        # 当前筛选账本（含备注列）
        self.tbl_ledger.setRowCount(0)
        try:
            rows = db.execute(
                "SELECT tx_id, created_at, tx_type, room_id, amount, note, pay_method, is_deposit "
                f"FROM ledger WHERE {today_filter} ORDER BY id DESC LIMIT 500"
            ).fetchall()
        except Exception:
            rows = []
        type_filter = self.cmb_ledger_type.currentData() if hasattr(self, "cmb_ledger_type") else "ALL"
        pay_filter = self.cmb_ledger_pay.currentData() if hasattr(self, "cmb_ledger_pay") else "ALL"
        needle = (self.txt_ledger_search.text() if hasattr(self, "txt_ledger_search") else "").strip().lower()
        filtered = []
        for row in rows:
            tx_id, ts, tx_type, room, amt, note, pay, is_dep = row
            if pay_filter != "ALL" and (pay or "") != pay_filter and normalize_pay_method(pay) != pay_filter:
                continue
            if type_filter == "REVENUE" and tx_type not in LEDGER_REVENUE_TX_TYPES:
                continue
            if type_filter == "DEPOSIT" and tx_type not in LEDGER_DEPOSIT_TX_TYPES:
                continue
            if type_filter == "PAYOUT" and tx_type != "PAYOUT":
                continue
            if type_filter == "RECONCILE" and tx_type != "CASH_RECONCILE":
                continue
            if str(type_filter).startswith("EXP:"):
                if tx_type != "PAYOUT" or _payout_category_from_note(note) != type_filter.split(":", 1)[1]:
                    continue
            if needle and needle not in " ".join(str(x or "") for x in row).lower():
                continue
            filtered.append(row)
        for i, row in enumerate(filtered[:200]):
            tx_id, ts, tx_type, room, amt, note, pay, is_dep = row
            self.tbl_ledger.insertRow(i)
            self.tbl_ledger.setItem(i, 0, QTableWidgetItem(str(ts or "")[:16]))
            self.tbl_ledger.setItem(i, 1, QTableWidgetItem(ledger_tx_type_display(tx_type, is_dep)))
            self.tbl_ledger.setItem(i, 2, QTableWidgetItem(pay_method_label(pay)))
            self.tbl_ledger.setItem(i, 3, QTableWidgetItem(str(room or "-")))
            amt_val = float(amt or 0)
            amt_item = QTableWidgetItem(f"{cur}{amt_val:.2f}")
            amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            amt_item.setForeground(QColor(_p("amount_positive") if amt_val >= 0 else _p("amount_negative")))
            self.tbl_ledger.setItem(i, 4, amt_item)
            note_item = QTableWidgetItem(str(note or "")[:80])
            note_item.setToolTip(str(note or ""))
            self.tbl_ledger.setItem(i, 5, note_item)
        self._sync_table_heights()

    def _sync_table_heights(self) -> None:
        fd_sync_table_height(self.tbl_ledger, min_rows=4, max_rows=14)
        fd_sync_table_height(self.tbl_operator_summary, min_rows=2, max_rows=8)
        fd_sync_table_height(self.tbl_cat, min_rows=2, max_rows=10)
        if hasattr(self, "tbl_reconcile"):
            fd_sync_table_height(
                self.tbl_reconcile,
                min_rows=len(PAYMENT_METHODS),
                max_rows=len(PAYMENT_METHODS),
            )

    def _populate_operator_summary(self) -> None:
        self.tbl_operator_summary.setRowCount(0)
        try:
            rows = db.execute(
                """
                SELECT COALESCE(actor_id,'unknown') AS op,
                       COUNT(*) AS total_ops,
                       SUM(CASE WHEN reason='CARD_ISSUE' THEN 1 ELSE 0 END) AS card_issues,
                       SUM(CASE WHEN reason='CARD_CANCEL' THEN 1 ELSE 0 END) AS card_cancels,
                       SUM(CASE WHEN reason LIKE 'CHECKIN%' OR reason LIKE 'PAYMENT%' THEN 1 ELSE 0 END) AS checkin_pay,
                       MAX(created_at) AS last_at
                FROM audit_events
                WHERE event_type='USER_ACTION' AND date(created_at)=date('now','localtime')
                GROUP BY COALESCE(actor_id,'unknown')
                ORDER BY total_ops DESC, op
                """
            ).fetchall()
        except Exception:
            rows = []
        for i, row in enumerate(rows):
            op, total, issues, cancels, checkin_pay, last_at = row
            self.tbl_operator_summary.insertRow(i)
            op_item = QTableWidgetItem(str(op or "unknown"))
            op_item.setData(Qt.ItemDataRole.UserRole, str(op or "unknown"))
            self.tbl_operator_summary.setItem(i, 0, op_item)
            for col, val in [(1, total), (2, issues), (3, cancels), (4, checkin_pay)]:
                num_item = QTableWidgetItem(str(val or 0))
                num_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.tbl_operator_summary.setItem(i, col, num_item)
            self.tbl_operator_summary.setItem(i, 5, QTableWidgetItem(str(last_at or "")[:16]))

    def _show_operator_detail(self, item) -> None:
        row = item.row() if item else -1
        if row < 0:
            return
        op_item = self.tbl_operator_summary.item(row, 0)
        op = op_item.data(Qt.ItemDataRole.UserRole) if op_item else ""
        if not op:
            return
        try:
            rows = db.execute(
                """
                SELECT created_at, reason, metadata_json
                FROM audit_events
                WHERE event_type='USER_ACTION'
                  AND COALESCE(actor_id,'unknown')=?
                  AND date(created_at)=date('now','localtime')
                ORDER BY created_at DESC
                LIMIT 80
                """,
                (op,),
            ).fetchall()
        except Exception as exc:
            show_warning(self, i18n.t("dlg_operator_detail"), str(exc))
            return
        lines = [f"{r[0]}  {r[1]}  {str(r[2] or '')[:120]}" for r in rows]
        show_info(self, i18n.t("dlg_operator_detail").format(op=op), "\n".join(lines) or i18n.t("msg_no_detail_today"))

    def _do_payout(self):
        """完整支出登记对话框（含支付方式）"""
        d = QDialog(self); d.setWindowTitle(i18n.t("finance_register_expense_title")); style_dialog(d, size="small")
        lv = QVBoxLayout(d); lv.setContentsMargins(FD_MARGIN, FD_MARGIN, FD_MARGIN, FD_MARGIN); lv.setSpacing(FD_SPACE_SM)
        lv.addWidget(build_dialog_header(
            i18n.t("finance_register_expense_title"),
            i18n.t("finance_register_expense_sub"),
        ))
        f = QFormLayout()
        cmb_type = QComboBox()
        for code, msg_key in EXPENSE_TYPES:
            cmb_type.addItem(i18n.t(msg_key), code)
        spn_amt = QDoubleSpinBox(); spn_amt.setRange(0.01, 999999); spn_amt.setDecimals(2)
        spn_amt.setPrefix(i18n.t("currency_symbol")); spn_amt.setValue(0.01)
        cmb_pay = _finance_pay_methods_combo()
        txt_note = QLineEdit(); txt_note.setPlaceholderText(i18n.t("finance_note_optional_ph"))
        f.addRow(i18n.t("finance_field_expense_type"), cmb_type)
        f.addRow(i18n.t("finance_field_amount"), spn_amt)
        f.addRow(i18n.t("finance_field_pay_method"), cmb_pay)
        f.addRow(i18n.t("finance_field_note_detail"), txt_note)
        btn_ok = QPushButton(i18n.t("finance_confirm_register")); btn_ok.setObjectName("SolidPrimaryBtn"); btn_ok.clicked.connect(d.accept)
        f.addRow(btn_ok)
        lv.addLayout(f)
        if not d.exec():
            return
        amt = spn_amt.value()
        t_code = cmb_type.currentData() or "MISC"
        t_label = i18n.t(_EXPENSE_CODE_TO_KEY.get(t_code, "expense_misc"))
        pay_method = cmb_pay.currentData() or "CASH_USD"
        cur = i18n.t("currency_symbol")
        if amt >= 300 and not ask_confirm(self, i18n.t("dlg_large_expense_confirm"), i18n.t("msg_large_expense").format(cur=cur, amt=amt, label=t_label, pay=pay_method_label(pay_method))):
            return
        note_text = txt_note.text().strip()
        full_note = f"{t_code}:{note_text}" if note_text else str(t_code)
        try:
            db.append_ledger("PAYOUT", -amt, "CASH", 1, note=full_note, pay_method=pay_method)
            bus.show_success_overlay.emit(i18n.t("finance_expense_registered").format(cur, amt))
            self.refresh()
            try:
                from telegram_shadow import telegram_thread
                if telegram_thread.isRunning():
                    telegram_thread.notify_payout(t_label, amt, note_text)
            except Exception:
                pass
        except Exception as e:
            show_warning(self, i18n.t("finance_register_failed"), str(e))

    # ── [sub-d Task1] 历史账单查询入口 ──────────────────────────
    def _open_bill_history(self):
        """打开账单详情对话框（搜索模式：按房号 / 日期查 bill_headers）。"""
        try:
            from tabs.frontdesk.bill_detail_dialog import BillDetailDialog
        except Exception as e:
            show_warning(self, "历史账单", f"账单查询模块加载失败：{e}")
            return
        try:
            BillDetailDialog(self).exec()
        except Exception as e:
            show_warning(self, "历史账单", f"打开失败：{e}")

    # ── [sub-d Task3] 多币种对账 / 月报入口 ─────────────────────
    def _open_multi_reconcile(self):
        """打开多币种对账对话框。"""
        try:
            MultiReconcileDialog(self).exec()
        except Exception as e:
            show_warning(self, "多币种对账", f"打开失败：{e}")

    def _open_multi_report(self):
        """打开多币种月报对话框。"""
        try:
            MultiCurrencyReportDialog(self).exec()
        except Exception as e:
            show_warning(self, "多币种月报", f"打开失败：{e}")


# ═══════════════════════════════════════════════════════════
# [sub-d Task3] 多币种对账 / 多币种月报对话框
# ═══════════════════════════════════════════════════════════


class MultiReconcileDialog(QDialog):
    """[sub-d Task3] 多币种对账对话框。

    数据源：reconciliation_service.daily_reconcile_multi_currency(date_str)
    展示：每行一个币种（原币小计 / 记账汇率折本位币 / 对账日汇率折本位币 / 汇兑损益）
          + 本位币合计 + 汇兑损益合计 + 异常清单（diff_lines）
    导出：CSV
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("多币种对账")
        style_dialog(self, size="large")
        self._data: dict = {}

        root = QVBoxLayout(self)
        root.setContentsMargins(FD_MARGIN, FD_MARGIN, FD_MARGIN, FD_MARGIN)
        root.setSpacing(FD_SPACE_SM)
        root.addWidget(build_dialog_header(
            "多币种对账",
            "按 ledger.currency 分组 + 按 exchange_rate 折算本位币 + 汇兑损益",
        ))

        # ── 顶部日期选择 ──
        from PySide6.QtWidgets import QDateEdit, QFormLayout
        top_row = QHBoxLayout()
        top_row.setSpacing(FD_SPACE_SM)
        self.dt_date = QDateEdit()
        self.dt_date.setCalendarPopup(True)
        self.dt_date.setDisplayFormat("yyyy-MM-dd")
        self.dt_date.setDate(datetime.date.today())
        self.dt_date.setMaximumWidth(160)
        top_row.addWidget(QLabel("对账日期："))
        top_row.addWidget(self.dt_date)
        self.btn_run = QPushButton("🔍 对账")
        self.btn_run.setObjectName("SolidPrimaryBtn")
        self.btn_run.clicked.connect(self._run_reconcile)
        top_row.addWidget(self.btn_run)
        top_row.addStretch()
        self.btn_export = QPushButton("导出 CSV")
        self.btn_export.setObjectName("FdCardActionBtn")
        self.btn_export.clicked.connect(self._export_csv)
        top_row.addWidget(self.btn_export)
        root.addLayout(top_row)

        # ── 主表格：每行一个币种 ──
        self.tbl = QTableWidget(0, 5)
        self.tbl.setObjectName("MultiReconcileTable")
        self.tbl.setHorizontalHeaderLabels(
            ["币种", "原币小计", "记账汇率折本位币", "对账日汇率折本位币", "汇兑损益"]
        )
        th = self.tbl.horizontalHeader()
        th.setMinimumSectionSize(80)
        th.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        for c in (1, 2, 3, 4):
            th.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        root.addWidget(self.tbl)

        # ── 底部合计行 ──
        self.lbl_total = QLabel("本位币合计：-　|　汇兑损益合计：-")
        self.lbl_total.setObjectName("H4Title")
        root.addWidget(self.lbl_total)

        # ── 异常清单展开区 ──
        from PySide6.QtWidgets import QGroupBox
        gb_diff = QGroupBox("异常清单（汇率差异 / 缺失支付方式）")
        gb_diff_lay = QVBoxLayout(gb_diff)
        self.tbl_diff = QTableWidget(0, 6)
        self.tbl_diff.setObjectName("MultiReconcileDiffTable")
        self.tbl_diff.setHorizontalHeaderLabels(
            ["类型", "币种", "房号/Tx", "原币", "记账汇率", "对账汇率"]
        )
        dh = self.tbl_diff.horizontalHeader()
        dh.setMinimumSectionSize(60)
        dh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        for c in (1, 2, 3, 4, 5):
            dh.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        self.tbl_diff.verticalHeader().setVisible(False)
        self.tbl_diff.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        gb_diff_lay.addWidget(self.tbl_diff)
        root.addWidget(gb_diff)

        # 关闭按钮
        btn_close_row = QHBoxLayout()
        btn_close_row.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.setObjectName("FdGhostBtn")
        btn_close.clicked.connect(self.accept)
        btn_close_row.addWidget(btn_close)
        root.addLayout(btn_close_row)

        # 首次加载今日对账
        self._run_reconcile()

    def _fmt(self, v) -> str:
        try:
            return f"{float(v or 0):.2f}"
        except Exception:
            return "0.00"

    def _run_reconcile(self) -> None:
        date_str = self.dt_date.date().toString("yyyy-MM-dd")
        try:
            from reconciliation_service import daily_reconcile_multi_currency
            data = daily_reconcile_multi_currency(date_str)
        except Exception as e:
            show_warning(self, "对账失败", str(e))
            return
        self._data = data or {}
        cur_sym = i18n.t("currency_symbol")
        base_cur = self._data.get("base_currency", "USD")

        # 主表格
        self.tbl.setRowCount(0)
        for i, c in enumerate(self._data.get("by_currency", [])):
            self.tbl.insertRow(i)
            self.tbl.setItem(i, 0, QTableWidgetItem(str(c.get("currency", ""))))
            orig_item = QTableWidgetItem(self._fmt(c.get("orig_subtotal")))
            orig_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 1, orig_item)
            base_item = QTableWidgetItem(f"{cur_sym}{self._fmt(c.get('base_subtotal'))}")
            base_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 2, base_item)
            actual_item = QTableWidgetItem(f"{cur_sym}{self._fmt(c.get('actual_base'))}")
            actual_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 3, actual_item)
            gain_loss = float(c.get("exchange_gain_loss") or 0)
            gl_item = QTableWidgetItem(f"{cur_sym}{self._fmt(gain_loss)}")
            gl_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            # 正绿负红
            gl_item.setForeground(QColor(_p("amount_positive") if gain_loss >= 0 else _p("danger")))
            self.tbl.setItem(i, 4, gl_item)

        # 合计行
        grand_total = self._fmt(self._data.get("grand_total_base"))
        gl_total = float(self._data.get("exchange_gain_loss_total") or 0)
        gl_total_color = _p("amount_positive") if gl_total >= 0 else _p("danger")
        self.lbl_total.setText(
            f"本位币合计（{base_cur}）：{cur_sym}{grand_total}　|　"
            f"<span style='color:{gl_total_color}'>汇兑损益合计：{cur_sym}{self._fmt(gl_total)}</span>"
        )
        self.lbl_total.setTextFormat(Qt.TextFormat.RichText)

        # 异常清单
        self.tbl_diff.setRowCount(0)
        for i, d in enumerate(self._data.get("diff_lines", [])):
            self.tbl_diff.insertRow(i)
            kind = d.get("kind", "")
            kind_zh = "汇率差异" if kind == "exchange_rate_diff" else (
                "缺失支付方式" if kind == "missing_pay_method" else kind
            )
            self.tbl_diff.setItem(i, 0, QTableWidgetItem(kind_zh))
            self.tbl_diff.setItem(i, 1, QTableWidgetItem(str(d.get("currency", ""))))
            self.tbl_diff.setItem(i, 2, QTableWidgetItem(
                str(d.get("room_id") or d.get("ledger_id") or d.get("tx_type") or "")
            ))
            self.tbl_diff.setItem(i, 3, QTableWidgetItem(self._fmt(d.get("orig") or d.get("amount"))))
            self.tbl_diff.setItem(i, 4, QTableWidgetItem(self._fmt(d.get("recorded_rate"))))
            self.tbl_diff.setItem(i, 5, QTableWidgetItem(self._fmt(d.get("actual_rate"))))

    def _export_csv(self) -> None:
        if not self._data:
            show_warning(self, "导出", "无可导出数据，请先执行对账。")
            return
        try:
            from pathlib import Path
            import csv as _csv
            out_dir = Path(db.get_config("report_export_dir") or "reports")
            out_dir.mkdir(exist_ok=True)
            fname = f"multi_reconcile_{self.dt_date.date().toString('yyyyMMdd')}.csv"
            path = out_dir / fname
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = _csv.writer(f)
                w.writerow(["多币种对账", self._data.get("date", "")])
                w.writerow(["本位币", self._data.get("base_currency", "")])
                w.writerow([])
                w.writerow(["币种", "原币小计", "记账汇率折本位币", "对账日汇率折本位币", "汇兑损益"])
                for c in self._data.get("by_currency", []):
                    w.writerow([
                        c.get("currency", ""),
                        self._fmt(c.get("orig_subtotal")),
                        self._fmt(c.get("base_subtotal")),
                        self._fmt(c.get("actual_base")),
                        self._fmt(c.get("exchange_gain_loss")),
                    ])
                w.writerow([])
                w.writerow(["本位币合计", self._fmt(self._data.get("grand_total_base"))])
                w.writerow(["汇兑损益合计", self._fmt(self._data.get("exchange_gain_loss_total"))])
                w.writerow([])
                w.writerow(["异常清单"])
                w.writerow(["类型", "币种", "房号/Tx", "原币", "记账汇率", "对账汇率"])
                for d in self._data.get("diff_lines", []):
                    w.writerow([
                        d.get("kind", ""),
                        d.get("currency", ""),
                        d.get("room_id") or d.get("ledger_id") or d.get("tx_type") or "",
                        self._fmt(d.get("orig") or d.get("amount")),
                        self._fmt(d.get("recorded_rate")),
                        self._fmt(d.get("actual_rate")),
                    ])
            show_info(self, "导出完成", f"已导出：{path}")
        except Exception as e:
            show_warning(self, "导出失败", str(e))


class MultiCurrencyReportDialog(QDialog):
    """[sub-d Task3] 多币种月报对话框。

    数据源：report_engine.ReportData.multi_currency_summary(year, month)
    展示：每行一个币种（原币总额 / 平均汇率 / 本位币合计）+ 本位币总计
    导出：Excel（openpyxl，已存在于项目）
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("多币种月报")
        style_dialog(self, size="large")
        self._data: dict = {}

        root = QVBoxLayout(self)
        root.setContentsMargins(FD_MARGIN, FD_MARGIN, FD_MARGIN, FD_MARGIN)
        root.setSpacing(FD_SPACE_SM)
        root.addWidget(build_dialog_header(
            "多币种月报",
            "按 ledger.currency 分组 + AVG(exchange_rate) 折本位币合计",
        ))

        # ── 顶部年月选择 ──
        from PySide6.QtWidgets import QSpinBox, QFormLayout
        now = datetime.date.today()
        top_row = QHBoxLayout()
        top_row.setSpacing(FD_SPACE_SM)
        top_row.addWidget(QLabel("年："))
        self.spn_year = QSpinBox()
        self.spn_year.setRange(2020, 2100)
        self.spn_year.setValue(now.year)
        self.spn_year.setMaximumWidth(100)
        top_row.addWidget(self.spn_year)
        top_row.addWidget(QLabel("月："))
        self.spn_month = QSpinBox()
        self.spn_month.setRange(1, 12)
        self.spn_month.setValue(now.month)
        self.spn_month.setMaximumWidth(60)
        top_row.addWidget(self.spn_month)
        self.btn_run = QPushButton("🔍 查询")
        self.btn_run.setObjectName("SolidPrimaryBtn")
        self.btn_run.clicked.connect(self._run_report)
        top_row.addWidget(self.btn_run)
        top_row.addStretch()
        self.btn_export = QPushButton("导出 Excel")
        self.btn_export.setObjectName("FdCardActionBtn")
        self.btn_export.clicked.connect(self._export_excel)
        top_row.addWidget(self.btn_export)
        root.addLayout(top_row)

        # ── 主表格 ──
        self.tbl = QTableWidget(0, 4)
        self.tbl.setObjectName("MultiCurrencyReportTable")
        self.tbl.setHorizontalHeaderLabels(
            ["币种", "原币总额", "平均汇率", "本位币合计"]
        )
        th = self.tbl.horizontalHeader()
        th.setMinimumSectionSize(80)
        th.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        for c in (1, 2, 3):
            th.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        root.addWidget(self.tbl)

        # ── 底部总计 ──
        self.lbl_total = QLabel("本位币总计：-")
        self.lbl_total.setObjectName("H4Title")
        root.addWidget(self.lbl_total)

        # 关闭按钮
        btn_close_row = QHBoxLayout()
        btn_close_row.addStretch()
        btn_close = QPushButton("关闭")
        btn_close.setObjectName("FdGhostBtn")
        btn_close.clicked.connect(self.accept)
        btn_close_row.addWidget(btn_close)
        root.addLayout(btn_close_row)

        # 首次加载本月
        self._run_report()

    def _fmt(self, v) -> str:
        try:
            return f"{float(v or 0):.2f}"
        except Exception:
            return "0.00"

    def _run_report(self) -> None:
        year = int(self.spn_year.value())
        month = int(self.spn_month.value())
        try:
            from report_engine import ReportData
            data = ReportData.multi_currency_summary(year, month)
        except Exception as e:
            show_warning(self, "查询失败", str(e))
            return
        self._data = data or {}
        cur_sym = i18n.t("currency_symbol")
        base_cur = self._data.get("base_currency", "USD")

        self.tbl.setRowCount(0)
        for i, c in enumerate(self._data.get("by_currency", [])):
            self.tbl.insertRow(i)
            self.tbl.setItem(i, 0, QTableWidgetItem(str(c.get("currency", ""))))
            orig_item = QTableWidgetItem(self._fmt(c.get("orig_total")))
            orig_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 1, orig_item)
            rate_item = QTableWidgetItem(self._fmt(c.get("avg_rate")))
            rate_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 2, rate_item)
            base_item = QTableWidgetItem(f"{cur_sym}{self._fmt(c.get('base_total'))}")
            base_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.tbl.setItem(i, 3, base_item)

        grand = self._fmt(self._data.get("grand_total_base"))
        self.lbl_total.setText(f"本位币总计（{base_cur}）：{cur_sym}{grand}")

    def _export_excel(self) -> None:
        if not self._data:
            show_warning(self, "导出", "无可导出数据，请先查询。")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from pathlib import Path
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "多币种月报"
            # 标题
            ws.merge_cells("A1:D1")
            ws["A1"] = f"{self._data.get('period', '')} 多币种月报（本位币 {self._data.get('base_currency', '')}）"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            # 表头
            headers = ["币种", "原币总额", "平均汇率", "本位币合计"]
            head_fill = PatternFill("solid", fgColor="1E3A5F")
            head_font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal="center")
            # 数据
            for i, c in enumerate(self._data.get("by_currency", []), start=3):
                ws.cell(row=i, column=1, value=str(c.get("currency", "")))
                ws.cell(row=i, column=2, value=float(c.get("orig_total") or 0))
                ws.cell(row=i, column=3, value=float(c.get("avg_rate") or 0))
                ws.cell(row=i, column=4, value=float(c.get("base_total") or 0))
            # 合计
            total_row = len(self._data.get("by_currency", [])) + 3
            ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=float(self._data.get("grand_total_base") or 0)).font = Font(bold=True)
            for col in range(1, 5):
                ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F2F0EE")
            # 列宽
            for col, w in enumerate([16, 18, 14, 18], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
            out_dir = Path(db.get_config("report_export_dir") or "reports")
            out_dir.mkdir(exist_ok=True)
            fname = f"multi_currency_{self._data.get('period', '').replace('-', '')}.xlsx"
            path = out_dir / fname
            wb.save(path)
            show_info(self, "导出完成", f"已导出：{path}")
        except Exception as e:
            show_warning(self, "导出失败", str(e))
